import re
import datetime as dt
import time
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from threading import Thread

def shift(shift_cell):
    time_regex = r"((([0]{1})?([1-9]{1})\:([0-9]{2}))|(([1-2]{1}[0-9]{1})\:([0-9]{2})))\s?-\s?((([0]{1})?([1-9]{1})\:([0-9]{2}))|((([1-2]{1})([0-9]{1}))\:([0-9]{2})))"
    shift_test = re.search(time_regex, shift_cell)
    shift_start = ""
    shift_end = ""
    if shift_test is not None:
        if shift_test.group(4) is not None: 
            shift_start = f"{shift_test.group(4)}:{shift_test.group(5)}"
        else:
            shift_start = shift_test.group(1)
        
        if shift_test.group(10) is not None: 
            shift_end = f"{shift_test.group(12)}:{shift_test.group(13)}"
        else:
            shift_end = shift_test.group(9)
        #print(shift_start)
        
    return (shift_start , shift_end, shift_test.group(5), shift_test.group(13))
    
def time_shift(cell):
    time_regex = r"(\d{1,2})[\-\:]{1}(\d{1,2})"
    time_test = re.search(time_regex, cell)
    time_hour = ""
    time_min = ""
    if time_test is not None:
        time_hour = time_test.group(1)
        time_min = time_test.group(2)
    return (time_hour, time_min)


def main() -> object:
    wb_grafik_per = openpyxl.load_workbook(f"перерывы_сборка.xlsx")
    for s in range(len(wb_grafik_per.sheetnames)):
        if wb_grafik_per.sheetnames[s] == 'перерывы':
            break
    wb_grafik_per.active = s

    sheet_per = wb_grafik_per.active



    #time_now = "15:23"
    #print(time_shift(time_now)[0],time_shift(time_now)[1])
    now = dt.datetime.now()
    #print(now)
    now_str = dt.datetime.strftime(now, "%H:%M")
    #print(now_str)
    time_now = now_str #f"{int(time_shift(now_str)[0])+12}:{time_shift(now_str)[1]}"
    #print(time_now)
    print(f"Сейчас в {time_now} на перерыве должны быть:")
    for m in range (10, 273, 12):
        if int(time_shift(shift(sheet_per.cell(row = m, column = 1).value)[0])[0]) <= int(time_shift(time_now)[0]) < int(time_shift(shift(sheet_per.cell(row = m, column = 1).value)[1])[0]):
            h_i = m
            #print(h_i)
            break
    for string in range(h_i, h_i + 12):
        if int(time_shift(sheet_per.cell(row = string, column = 2).value)[0]) <= int(time_shift(time_now)[1]) < int(time_shift(sheet_per.cell(row = string, column = 2).value)[1]):
            m_i = string
            #print(m_i)
            break
    for kol in range(4,sheet_per.max_column + 1 ):
        if sheet_per.cell(row = m_i, column = kol).value is not None:
            #print(sheet_per.cell(row = 2, column = kol).value)
            per_end = 10
            for st_per in range(m_i, sheet_per.max_row + 1):
                if sheet_per.cell(row = st_per, column = kol).value != 5:
                    per_end = st_per - 1
                    break
            for st_end in range(per_end, per_end - 12, -1):
                if sheet_per.cell(row = st_end, column = 1).value is not None:
                    hour_end = time_shift(shift(sheet_per.cell(row = st_end, column = 1).value)[0])[0]
                    break
            min_end = time_shift(sheet_per.cell(row = per_end, column = 2).value)[1]
            if min_end == "60":
                min_end = "00"
                hour_end = int(hour_end) + 1
            elif min_end == "0":
                min_end = "00"
            elif min_end == "5":
                min_end = "05"
    # ищем с которого перерыв
            for st_per_beg in range(per_end, 10, -1):
                if sheet_per.cell(row = st_per_beg, column = kol).value != 5:
                    per_beg = st_per_beg + 1
                    break
            for st_beg in range(per_beg, per_beg - 12, -1):
                if sheet_per.cell(row = st_beg, column = 1).value is not None:
                    hour_beg = time_shift(shift(sheet_per.cell(row = st_beg, column = 1).value)[0])[0]
                    break
            min_beg = time_shift(sheet_per.cell(row = per_beg, column = 2).value)[0]
            if min_beg == "60":
                min_beg = "00"
                hour_beg = int(hour_beg) + 1
            elif min_beg == "0":
                min_beg = "00"
            elif min_beg == "5":
                min_beg = "05"
            probel = " "
            #print(f"{sheet_per.cell(row = 2, column = kol).value} {probel*(20-len(sheet_per.cell(row = 2, column = kol).value))}{hour_beg}:{min_beg} - {hour_end}:{min_end}")
            #print(f"{sheet_per.cell(row = 2, column = kol).value} до {hour_end}:{min_end}")
            print(f"{sheet_per.cell(row=1, column=kol).value} "
                  f"{probel * (15 - len(sheet_per.cell(row=1, column=kol).value))}"
                  f"{sheet_per.cell(row=2, column=kol).value} "
                  f"{probel * (20 - len(sheet_per.cell(row=2, column=kol).value))}"
                  f"{hour_beg}:{min_beg} - {hour_end}:{min_end}")

    while True:
        now = dt.datetime.now()
        #print(now)
        now_str = dt.datetime.strftime(now, "%H:%M")
        #print(now_str)
        now_minus_6 = now_str #f"{int(time_shift(now_str)[0])+12}:{time_shift(now_str)[1]}"
        if not (int(time_shift(sheet_per.cell(row = m_i, column = 2).value)[0]) <= int(time_shift(now_minus_6)[1]) < int(time_shift(sheet_per.cell(row = m_i, column = 2).value)[1])):
            print("----------------------------------")
            print(f"На перерыве в {now_minus_6} должны быть: ")
            m_i += 1
            for kol in range(4,sheet_per.max_column + 1 ):
                if sheet_per.cell(row = m_i, column = kol).value is not None:
            #print(sheet_per.cell(row = 2, column = kol).value)
                    per_end = 10
                    for st_per in range(m_i, sheet_per.max_row + 1):
                        if sheet_per.cell(row = st_per, column = kol).value != 5:
                            per_end = st_per - 1
                            break
                    for st_end in range(per_end, per_end - 12, -1):
                        if sheet_per.cell(row = st_end, column = 1).value is not None:
                            hour_end = time_shift(shift(sheet_per.cell(row = st_end, column = 1).value)[0])[0]
                            break
                    min_end = time_shift(sheet_per.cell(row = per_end, column = 2).value)[1]
                    if min_end == "60":
                        min_end = "00"
                        hour_end = int(hour_end) + 1
                    elif min_end == "0":
                        min_end = "00"
                    elif min_end == "5":
                        min_end = "05"
                     # ищем с которого перерыв
                    for st_per_beg in range(per_end, 10, -1):
                        if sheet_per.cell(row = st_per_beg, column = kol).value != 5:
                            per_beg = st_per_beg + 1
                            break
                    for st_beg in range(per_beg, per_beg - 12, -1):
                        if sheet_per.cell(row = st_beg, column = 1).value is not None:
                            hour_beg = time_shift(shift(sheet_per.cell(row = st_beg, column = 1).value)[0])[0]
                            break
                    min_beg = time_shift(sheet_per.cell(row = per_beg, column = 2).value)[0]
                    if min_beg == "60":
                        min_beg = "00"
                        hour_beg = int(hour_beg) + 1
                    elif min_beg == "0":
                        min_beg = "00"
                    elif min_beg == "5":
                        min_beg = "05"
                    probel = " "
                    #print(f"{sheet_per.cell(row = 2, column = kol).value} {probel*(20-len(sheet_per.cell(row = 2, column = kol).value))}{hour_beg}:{min_beg} - {hour_end}:{min_end}")
                    print(f"{sheet_per.cell(row=1, column=kol).value} "
                        f"{probel * (15 - len(sheet_per.cell(row=1, column=kol).value))}"
                        f"{sheet_per.cell(row=2, column=kol).value} "
                        f"{probel * (20 - len(sheet_per.cell(row=2, column=kol).value))}"
                        f"{hour_beg}:{min_beg} - {hour_end}:{min_end}")
        time.sleep(5)
#    Thread(target=main).start()

if __name__ == "__main__":
    Thread(target=main).start()
