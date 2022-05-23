import re
import datetime as dt
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from tqdm import tqdm


def fio(fio_cell):
    # преобразование ФИО в Ф И.О.
    if fio_cell is None:
        return ""
    else:
        fio_regex = r"([А-ЯЁ]{1}[а-яё]+)\s{1}(\(([А-ЯЁ]{1}[а-яё]+)\)\s{1})?([А-ЯЁ]{1}[а-яё]+)(\s{1}([А-ЯЁ]{1}[а-яё]+))?"
        fio_test = re.search(fio_regex, fio_cell)
        fio_new = ""
        if fio_test is not None:
            if fio_test.group(6) is not None:
                fio_new = fio_test.group(1) + " " + fio_test.group(4)[0] + ". " + fio_test.group(6)[0] + "."
            else:
                fio_new = fio_test.group(1) + " " + fio_test.group(4)[0] + "."
            # print(fio_new)
    return fio_new


def oper_div(sheet_name):  # поиск в столбце с ФИО разделителей по сменам
    for i in range(1, sheet_name.max_row + 1):
        if sheet_name.cell(row=i, column=6).value == "ОПЕРАТОРЫ 5/2":
            i_5_2 = i
        # print(i_5_2)
        if sheet_name.cell(row=i, column=6).value == "ОПЕРАТОРЫ 2/2":
            i_2_2 = i
        # print(i_2_2)
        if (sheet_name.cell(row=i, column=6).value == "НЕПОЛНЫЙ РАБОЧИЙ ДЕНЬ" or
                sheet_name.cell(row=i, column=6).value == "ОПЕРАТОРЫ  неполного рабочего  дня"):
            i_1_2 = i
        # print(i_1_2)
    return i_5_2, i_2_2, i_1_2


def fio_full(fio_cell):
    # преобразование ФИО в Ф И.О.
    if fio_cell is None:
        return ""
    else:
        fio_regex = r"([А-ЯЁ]{1}[а-яё]+)\s{1}(\(([А-ЯЁ]{1}[а-яё]+)\)\s{1})?([А-ЯЁ]{1}[а-яё]+)(\s{1}([А-ЯЁ]{1}[а-яё]+))?"
        fio_test = re.search(fio_regex, fio_cell)
        fio_new = ""
        if fio_test is not None:
            # print(fio_test.group(0))
            # print(fio_test.group(1))
            # print(fio_test.group(2))
            # print(fio_test.group(3))
            # print(fio_test.group(4))

            if fio_test.group(6) is not None:
                fio_new = fio_test.group(0)
            else:
                fio_new = fio_test.group(1) + " " + fio_test.group(4)
            # print(fio_
    return fio_new


def shift(shift_cell):
    time_regex = r"((([0]{1})?([1-9]{1})\:([0-9]{2}))|(([1-2]{1}[0-9]{1})\:" \
                 r"([0-9]{2})))\s?-\s?((([0]{1})?([1-9]{1})\:([0-9]{2}))|((([1-2]{1})([0-9]{1}))\:([0-9]{2})))"
    shift_test = re.search(time_regex, shift_cell)
    shift_start = ""
    shift_end = ""
    if shift_test is not None:
        if shift_test.group(4) is not None:
            shift_start = f"{shift_test.group(4)}:{shift_test.group(5)}"
        else:
            shift_start = shift_test.group(1)

        if shift_test.group(10) is not None:
            shift_end = f"{shift_test.group(12)}:{shift_test.group(13)}"
        else:
            shift_end = shift_test.group(9)
        # print(shift_start)

    return shift_start, shift_end, shift_test.group(5), shift_test.group(13)


def time_shift(cell):
    time_regex = r"(\d{1,2})[\-\:]{1}(\d{1,2})"
    time_test = re.search(time_regex, cell)
    time_hour = ""
    time_min = ""
    if time_test is not None:
        time_hour = time_test.group(1)
        time_min = time_test.group(2)
    return time_hour, time_min


def sting_no(sheet):
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=6).value == "ОПЕРАТОРЫ 5/2":
            i_5_2 = i
            # print(i_5_2)
        if sheet.cell(row=i, column=6).value == "ОПЕРАТОРЫ 2/2":
            i_2_2 = i
            # print(i_2_2)
        if (sheet.cell(row=i, column=6).value == "НЕПОЛНЫЙ РАБОЧИЙ ДЕНЬ" or
                sheet.cell(row=i, column=6).value == "ОПЕРАТОРЫ  неполного рабочего  дня"):
            i_1_2 = i
            # print(i_1_2)
    return i_5_2, i_2_2, i_1_2


def find_cell(k, sheet, CC_name, shift):
    wb_per = openpyxl.load_workbook("Исключения.xlsx")
    sheet_per = wb_per.active

    # n1 = ""
    # n2 = ""
    # n3 = ""
    # n4 = ""
    # n5 = ""
    # n6 = 0
    # n7 = 0
    color = 0
    # oper = ""
    oper = fio_full(sheet.cell(row=k, column=6).value)
    # print(oper)
    n1 = "Смирновка"
    n2 = fio(oper)
    n3 = "5/2"
    # print(type(sheet.cell(row = k, column = 5).value))
    # temp = sheet.cell(row = k, column = 5).value
    # print(shift(temp)[0])
    # n4 = shift(sheet.cell(row = k, column = 5).value)[0]
    # n5 = shift(sheet.cell(row = k, column = 5).value)[1]
    cell = sheet.cell(row=k, column=9).value
    cell_plus_1 = sheet.cell(row=k + 1, column=9).value

    if (type(cell) == int or type(cell) == float) and (type(cell_plus_1) == int or type(cell_plus_1) == float):
        n4 = "с+доп"
    elif (type(cell) == int or type(cell) == float) and cell_plus_1 is None:
        n4 = "c"
    else:
        n4 = "доп"

    if (type(cell) == int or type(cell) == float) and (type(cell_plus_1) == int or type(cell_plus_1) == float):
        n6 = sheet.cell(row=k, column=9).value + sheet.cell(row=k + 1, column=9).value
    elif (type(cell) == int or type(cell) == float) and cell_plus_1 is None:
        n6 = sheet.cell(row=k, column=9).value
    else:
        n6 = sheet.cell(row=k + 1, column=9).value

    #     if sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is not None:
    #         n4 = "с+доп"
    #     elif sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is None:
    #         n4 = "c"
    #     else:
    #         n4 = "доп"

    #     if sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is not None:
    #         n6 = sheet.cell(row = k, column = 9).value + sheet.cell(row = k+1, column = 9).value
    #     elif sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is None:
    #         n6 = sheet.cell(row = k, column = 9).value
    #     else:
    #         n6 = sheet.cell(row = k+1, column = 9).value

    if n4 == "c":
        if n6 == 11:
            n7 = 75

        elif n6 > 11:
            n7 = 75
            color = 1
        elif n6 == 8:
            n7 = 65
        elif 8 < n6 < 11:
            n7 = 70
            color = 1
        elif n6 == 3:
            n7 = 15
            color = 1
        elif 3 < n6 < 6:
            n7 = 20
            color = 1
        elif n6 == 6:
            n7 = 30
            color = 1
        elif 6 < n6 < 8:
            n7 = 60
            color = 1
        else:
            n7 = ""
            color = 1

    else:
        if n6 == 11:
            n7 = 75
        elif 8 < n6 < 11:
            n7 = 70

        elif n6 > 11:
            n7 = 75

        elif n6 == 8:
            n7 = 65
        elif 3 < n6 < 6:
            n7 = 20

        elif n6 == 3:
            n7 = 15

        elif n6 == 6:
            n7 = 30

        elif 6 < n6 < 8:
            n7 = 60

        else:
            n7 = ""
        color = 1
    for i in range(2, sheet_per.max_row + 1):
        if n2 == sheet_per.cell(row=i, column=1).value:
            n7 = sheet_per.cell(row=i, column=2).value
            if n4 == "c":
                color = 0
            else:
                color = 1
    return n1, n2, n3, n4, n6, n7, oper, color


def sum_cell(line1, line2, sheet):
    sum_lines = 0
    for z in range(4, sheet.max_column + 1):
        for line in range(line1, line2):
            if sheet.cell(row=line, column=z).value is not None:
                sum_lines += sheet.cell(row=line, column=z).value
    return sum_lines


def shift_yellow(sheet_rez1, g1, time_start1, time_end1):
    i_shift_start = 10
    i_shift_end = 0
    for l in range(10, sheet_rez1.max_row + 1):
        # запоминаем строку с началом смены
        if sheet_rez1.cell(row=l, column=1).value is not None:
            if shift(sheet_rez1.cell(row=l, column=1).value)[0] == time_start1:
                i_shift_start = l
                # print(i_shift_start)
            elif int(time_shift(shift(sheet_rez1.cell(row=l, column=1).value)[0])[0]) <= int(
                    time_shift(time_start1)[0]) < int(time_shift(shift(sheet_rez1.cell(row=l, column=1).value)[1])[0]):
                for e in range(l, l + 12):
                    if int(time_shift(time_start1)[1]) == int(time_shift(sheet_rez1.cell(row=e, column=2).value)[0]):
                        i_shift_start = e
                        # print(i_shift_start)
        # запоминаем строку с концом смены
        if sheet_rez1.cell(row=l, column=1).value is not None:
            if shift(sheet_rez1.cell(row=l, column=1).value)[1] == time_end1:
                i_shift_end = l + 11
                # print(i_shift_end)
            elif int(time_shift(shift(sheet_rez1.cell(row=l, column=1).value)[0])[0]) <= int(
                    time_shift(time_end1)[0]) < int(time_shift(shift(sheet_rez1.cell(row=l, column=1).value)[1])[0]):
                for e in range(l, l + 12):
                    if int(time_shift(time_end1)[1]) == int(time_shift(sheet_rez1.cell(row=e, column=2).value)[1]):
                        i_shift_end = e

                    elif int(time_shift(time_end1)[1]) > int(time_shift(sheet_rez1.cell(row=e, column=2).value)[1]):
                        i_shift_end = e
                        # print(i_shift_end)
    # print(sheet_rez.cell(row=2, column=g).value)
    # print(i_shift_start)
    # print(i_shift_end)
    # закрашиваем смену
    yellowFill = PatternFill(start_color='FCF305', end_color='FCF305', fill_type='solid')
    for z in range(i_shift_start, i_shift_end + 1):
        sheet_rez1.cell(row=z, column=g1).fill = yellowFill
        sheet_rez1.cell(row=z, column=g1).value = 5
    return sheet_rez1


def set_pereriv(sheet_rez, len_per_set, time_per_set, i_shift_start_set, i_shift_end_set, i_p=0, n_p=0, prnt=0):
    if len_per_set > 1:
        break_time = int((i_shift_end_set + 1 - i_shift_start_set) / (len_per_set + 1))

        #i_p = 0
        for w in range(i_shift_start_set + break_time, i_shift_end_set - pereriv[time_per_set][-1], break_time):
            break_line = w
            break_sum = sum_cell(w, w + pereriv[time_per_set][i_p], sheet_rez)
            for break_lines in range(w - 6, w + 6 + 1):  # - pereriv[time_per][i_p]
                if break_sum > sum_cell(break_lines, break_lines + pereriv[time_per_set][i_p], sheet_rez):
                    break_line = break_lines
                    break_sum = sum_cell(break_lines, break_lines + pereriv[time_per_set][i_p], sheet_rez)
                elif break_sum == sum_cell(break_lines, break_lines + pereriv[time_per_set][i_p], sheet_rez) and abs(
                        w - break_lines) < abs(w - break_line):
                    break_line = break_lines
                    break_sum = sum_cell(break_lines, break_lines + pereriv[time_per_set][i_p], sheet_rez)
            # print(break_line)
            for ii_p in range(pereriv[time_per_set][i_p]):
                # print(ii_p)
                sheet_rez.cell(row=break_line + ii_p, column=g).value = 5
                sheet_rez.cell(row=break_line + ii_p, column=g).fill = yellowFill
            i_p += 1
            if i_p == n_p:
                i_p += 1
    elif len_per_set == 1:
        break_time = int((i_shift_end_set + 1 - i_shift_start_set) / 2)
        if type(pereriv[time_per_set]) == list:
            range_i = pereriv[time_per_set][i_p]
        else:
            range_i = pereriv[time_per_set]
        for ij_p in range(range_i):
            sheet_rez.cell(row=i_shift_start_set + break_time + ij_p, column=g).value = 5
            sheet_rez.cell(row=i_shift_start_set + break_time + ij_p, column=g).fill = yellowFill
        i_p += 1
        if i_p == n_p:
            i_p += 1
    return sheet_rez, i_p

def posledniy(sheet_rez, len_per, time_per, i_shift_start, i_shift_end):
    for ii_p in range(pereriv[time_per][-1]):
        # print(ii_p)
        sheet_rez.cell(row=i_shift_end - ii_p, column=g).value = 5
        sheet_rez.cell(row=i_shift_end - ii_p, column=g).fill = yellowFill
    i_shift_end -= ii_p

    if len_per > 1:
        break_time = int((i_shift_end + 1 - i_shift_start) / (len_per))

        i_p = 0
        for w in range(i_shift_start + break_time, i_shift_end - pereriv[time_per][-2], break_time):
            break_line = w
            break_sum = sum_cell(w, w + pereriv[time_per][i_p], sheet_rez)
            for break_lines in range(w - 6, w + 6 + 1):  # - pereriv[time_per][i_p]
                if break_sum > sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez):
                    break_line = break_lines
                    break_sum = sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez)
                elif break_sum == sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez) and abs(
                        w - break_lines) < abs(w - break_line):
                    break_line = break_lines
                    break_sum = sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez)
            # print(break_line)
            for ii_p in range(pereriv[time_per][i_p]):
                # print(ii_p)
                sheet_rez.cell(row=break_line + ii_p, column=g).value = 5
                sheet_rez.cell(row=break_line + ii_p, column=g).fill = yellowFill
            i_p += 1

    return sheet_rez


def perviy(sheet_rez, len_per, time_per, i_shift_start, i_shift_end, sheet_fix, i_fix):
    start = shift(sheet_fix.cell(row=i_fix, column=4).value)[0]
    stop = shift(sheet_fix.cell(row=i_fix, column=4).value)[1]
    i_start = nachalo_konec(sheet_rez, start, stop)[0]
    i_end = nachalo_konec(sheet_rez, start, stop)[1]
    #print(start, stop, i_start, i_end)
    for ii_p in range(pereriv[time_per][0]):
        # print(ii_p)
        sheet_rez.cell(row=i_start + ii_p, column=g).value = 5
        sheet_rez.cell(row=i_start + ii_p, column=g).fill = yellowFill
    #sheet_rez = shift_yellow(sheet_rez, g, i_shift_start, i_shift_end)


    if len_per > 1:
        break_time = int((i_shift_end + 1 - i_end) / (len_per))

    i_p = 1
    #print(pereriv[time_per])
    for w in range(i_end + break_time, i_shift_end - pereriv[time_per][-1], break_time):
        break_line = w
        break_sum = sum_cell(w, w + pereriv[time_per][i_p], sheet_rez)
        for break_lines in range(w - 6, w + 6 + 1):  # - pereriv[time_per][i_p]
            if break_sum > sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez):
                break_line = break_lines
                break_sum = sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez)
            elif break_sum == sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez) and abs(
                    w - break_lines) < abs(w - break_line):
                break_line = break_lines
                break_sum = sum_cell(break_lines, break_lines + pereriv[time_per][i_p], sheet_rez)
        # print(break_line)
        for ii_p in range(pereriv[time_per][i_p]):
            # print(ii_p)
            sheet_rez.cell(row=break_line + ii_p, column=g).value = 5
            sheet_rez.cell(row=break_line + ii_p, column=g).fill = yellowFill
        i_p += 1

    return sheet_rez


def nachalo_konec(sheet_rez, time_start, time_end):
    # закрашиваем шифты
    i_shift_start = 10
    i_shift_end = 0
    for l in range(10, sheet_rez.max_row + 1):
        # запоминаем строку с началом смены
        if sheet_rez.cell(row=l, column=1).value is not None:
            if shift(sheet_rez.cell(row=l, column=1).value)[0] == time_start:
                i_shift_start = l
                # print(i_shift_start)
            elif int(time_shift(shift(sheet_rez.cell(row=l, column=1).value)[0])[0]) <= int(
                    time_shift(time_start)[0]) < int(
                time_shift(shift(sheet_rez.cell(row=l, column=1).value)[1])[0]):
                for e in range(l, l + 12):
                    if time_shift(time_start)[1] == time_shift(sheet_rez.cell(row=e, column=2).value)[0]:
                        i_shift_start = e
                        # print(i_shift_start)
        # запоминаем строку с концом смены
        if sheet_rez.cell(row=l, column=1).value is not None:
            if shift(sheet_rez.cell(row=l, column=1).value)[1] == time_end:
                i_shift_end = l + 11
                # print(i_shift_end)
            elif int(time_shift(shift(sheet_rez.cell(row=l, column=1).value)[0])[0]) <= int(
                    time_shift(time_end)[0]) < int(time_shift(shift(sheet_rez.cell(row=l, column=1).value)[1])[0]):
                for e in range(l, l + 12):
                    if time_shift(time_end)[1] == time_shift(sheet_rez.cell(row=e, column=2).value)[1]:
                        i_shift_end = e

                    elif time_shift(time_end)[1] > time_shift(sheet_rez.cell(row=e, column=2).value)[1]:
                        i_shift_end = e
    return i_shift_start, i_shift_end



wb_rez = openpyxl.load_workbook("перерывы_сборка.xlsx")
for s in range(len(wb_rez.sheetnames)):
    if wb_rez.sheetnames[s] == 'перерывы':
        wb_rez.active = s
        break

sheet_rez = wb_rez.active
sheet_rez.delete_rows(10, 273)

redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
greenFill = PatternFill(start_color='1FB714', end_color='1FB714', fill_type='solid')
yellowFill = PatternFill(start_color='FCF305', end_color='FCF305', fill_type='solid')

#"""
# считываем файл фикс
wb_fix = openpyxl.load_workbook(f"Фикс.xlsx")
wb_fix.active = wb_fix['Фикс']
sheet_fix = wb_fix.active
#"""


# создавем словарь перерывов из файла
wb_pereriv = openpyxl.load_workbook(f"Словарь_перерывов.xlsx")
wb_pereriv.active = wb_pereriv['Основные']
sheet_pereriv = wb_pereriv.active

perrr = {}
for l_p in range(2, sheet_pereriv.max_row + 1):
    if sheet_pereriv.cell(row=l_p, column=1).value is not None:
        # print(str(sheet_pereriv.cell(row = l_p, column = 1).value))
        if sheet_pereriv.cell(row=l_p, column=3).value is None:
            perrr[str(sheet_pereriv.cell(row=l_p, column=1).value)] = int(
                int(sheet_pereriv.cell(row=l_p, column=2).value) / 5)
        else:
            list_per = []
            for r_p in range(2, sheet_pereriv.max_column + 2):
                if sheet_pereriv.cell(row=l_p, column=r_p).value is not None:
                    list_per.append(int(int(sheet_pereriv.cell(row=l_p, column=r_p).value) / 5))
                    # print(list_per)
                else:
                    perrr[str(sheet_pereriv.cell(row=l_p, column=1).value)] = list_per
                    # print(perrr)
                    break

wb_pereriv.active = wb_pereriv['Собственные']
sheet_pereriv = wb_pereriv.active

for l_p in range(2, sheet_pereriv.max_row + 1):
    if sheet_pereriv.cell(row=l_p, column=1).value is not None:
        # print(str(sheet_pereriv.cell(row = l_p, column = 1).value))
        if sheet_pereriv.cell(row=l_p, column=3).value is None:
            perrr[str(sheet_pereriv.cell(row=l_p, column=1).value)] = int(
                int(sheet_pereriv.cell(row=l_p, column=2).value) / 5)
        else:
            list_per = []
            for r_p in range(2, sheet_pereriv.max_column + 2):
                if sheet_pereriv.cell(row=l_p, column=r_p).value is not None:
                    list_per.append(int(int(sheet_pereriv.cell(row=l_p, column=r_p).value) / 5))
                    # print(list_per)
                else:
                    perrr[str(sheet_pereriv.cell(row=l_p, column=1).value)] = list_per
                    # print(perrr)
                    break
# print(perrr)

# создаем словарь с перерывами
# pereriv = {
#     "15" : 3,
#     "20" : [2,2],
#     "30" : [3,3],
#     "55" : [3,6,2],
#     "60" : [3,6,3],
#     "65" : [2,2,6,3],
#     "70" : [2,3,6,3],
#     "75" : [3,3,6,3],
#     "к65" :[2,3,2,3,3],
#     "к75" :[3,3,3,3,3]
# }
pereriv = perrr
# print(pereriv)
# заполняем таблицу слева
# sheet_rez.merge_cells('A10:A14')
# sheet_rez['A10'] = '6:30 - 7:00'
# count_hour = 7
# sheet_rez['B10'] = '30-35'
# sheet_rez['B11'] = '35-40'
# sheet_rez['B12'] = '40-45'
# sheet_rez['B13'] = '45-50'
# sheet_rez['B14'] = '50-55'

# заполняем первые два столбца
count_hour = 2
for m in range(10, 264, 12):
    sheet_rez.merge_cells(f"A{m}:A{m + 11}")
    sheet_rez[f"A{m}"] = f"{count_hour}:00 - {count_hour + 1}:00"
    count_hour += 1
    for i in range(0, 60, 5):
        sheet_rez[f"B{int(m + i / 5)}"] = f"{i}-{i + 5}"

# считаем в третьем столбце количество перерывов в интервале
for i in range(10, sheet_rez.max_row + 1):
    sheet_rez[f"C{i}"] = f"=COUNT(D{i}:{get_column_letter(sheet_rez.max_column)}{i})"

# рисуем границы
print("Подготовка таблицы....")
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
medium = Side(border_style="medium", color="000000")

for i in range(1, sheet_rez.max_column + 1):
    for j in range(1, 9):
        sheet_rez.cell(row=j, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for z in range(10, sheet_rez.max_row + 1):
        if sheet_rez.cell(row=z, column=1).value is None:
            sheet_rez.cell(row=z, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        else:
            sheet_rez.cell(row=z, column=i).border = Border(top=medium, left=thin, right=thin, bottom=thin)

# счтываем время начала, окончания и время перерывов и расставляем время
try:
    for g in tqdm(range(4, sheet_rez.max_column + 1), desc="Расставляем перерывы: ", unit=" ФИО", dynamic_ncols=True):
        time_per = str(sheet_rez.cell(row=8, column=g).value)
        # print(time_per)
        # проверяем ячейки со временем, которые полсе изменения могли поменять тип данных
        #     if type(sheet_rez.cell(row = 4, column = g).value) == dt.time:
        #         time_start = sheet_rez.cell(row = 4, column = g).value.strftime("%H:%M")
        #         if time_start[0] == "0":
        #             time_start = time_start[1:]
        #     else:
        #         time_start = str(sheet_rez.cell(row = 4, column = g).value)
        if type(sheet_rez.cell(row=4, column=g).value) == dt.time:
            sheet_rez.cell(row=4, column=g).value = sheet_rez.cell(row=4, column=g).value.strftime("%H:%M")
        if sheet_rez.cell(row=4, column=g).value[0] == "0":
            sheet_rez.cell(row=4, column=g).value = sheet_rez.cell(row=4, column=g).value[1:]
        if sheet_rez.cell(row=4, column=g).value[-1:] != "0" or sheet_rez.cell(row=4, column=g).value[-1:] != "5":
            temp_range = str(round(int(sheet_rez.cell(column=g, row=4).value[-2:]) / 10) * 10)
            if temp_range in range(10, 51):
                sheet_rez.cell(row=4, column=g).value = sheet_rez.cell(row=4, column=g).value[:-2] + temp_range
            elif int(temp_range) == 60:
                sheet_rez.cell(row=4, column=g).value = f'{int(sheet_rez.cell(row=4, column=g).value[:-3]) + 1}:00'
            elif int(temp_range) == 0:
                sheet_rez.cell(row=4, column=g).value = f'{sheet_rez.cell(row=4, column=g).value[:-2]}0{temp_range}'

            # if sheet_rez.cell(row=4, column=g).value[0] == "0":
            #     sheet_rez.cell(row=4, column=g).value = sheet_rez.cell(row=4, column=g).value[1:]
            # if sheet_rez.cell(row=4, column=g).value[:-1] != "0" or sheet_rez.cell(row=4, column=g).value[-1:] != "5":
            #     if round(int(sheet_rez.cell(column=g, row=4).value[-2:]) / 10) * 10 in range(1, 60):
            #         sheet_rez.cell(row=4, column=g).value = sheet_rez.cell(row=4, column=g).value[:-2] + str(
            #             round(int(sheet_rez.cell(column=g, row=4).value[-2:]) / 10) * 10)
            #     elif round(int(sheet_rez.cell(column=g, row=4).value[-2:]) / 10) * 10 in range(0, 6):
            #         sheet_rez.cell(row=4, column=g).value = sheet_rez.cell(row=4, column=g).value[:-2] + \
            #                                                 str(round(
            #                                                     int(sheet_rez.cell(column=g, row=4).value[-2:]) / 10) * 10) \
            #                                                 + "0"
            #     else:
            #         sheet_rez.cell(row=4, column=g).value = str(
            #             int(time_shift(sheet_rez.cell(row=4, column=g).value)[0]) + 1) + ":00"
        time_start = str(sheet_rez.cell(row=4, column=g).value)
        # print(type(time_start))
        # print(time_start)
        if time_start is None:
            break
        # print(time_start)
        #     if type(sheet_rez.cell(row = 5, column = g).value) == dt.time:
        #         time_end = sheet_rez.cell(row = 5, column = g).value.strftime("%H:%M")
        #         if time_end[0] == "0":
        #             time_end = time_end[1:]
        #     else:
        #         time_end = str(sheet_rez.cell(row = 5, column = g).value)
        if type(sheet_rez.cell(row=5, column=g).value) == dt.time:
            sheet_rez.cell(row=5, column=g).value = sheet_rez.cell(row=5, column=g).value.strftime("%H:%M")
        if sheet_rez.cell(row=5, column=g).value[0] == "0":
            sheet_rez.cell(row=5, column=g).value = sheet_rez.cell(row=5, column=g).value[1:]
        if sheet_rez.cell(row=5, column=g).value[-1:] != "0" or sheet_rez.cell(row=5, column=g).value[-1:] != "5":
            temp_range = str(round(int(sheet_rez.cell(column=g, row=5).value[-2:]) / 10) * 10)

            if int(temp_range) in range(10, 51):
                sheet_rez.cell(row=5, column=g).value = sheet_rez.cell(row=5, column=g).value[:-2] + temp_range
            elif int(temp_range) == 60:
                sheet_rez.cell(row=5, column=g).value = f'{int(sheet_rez.cell(row=5, column=g).value[:-3]) + 1}:00'
            elif int(temp_range) == 0:
                sheet_rez.cell(row=5, column=g).value = f'{sheet_rez.cell(row=5, column=g).value[:-2]}0{temp_range}'
            # if sheet_rez.cell(row=5, column=g).value[0] == "0":
            #     sheet_rez.cell(row=5, column=g).value = sheet_rez.cell(row=5, column=g).value[1:]
            # if sheet_rez.cell(row=5, column=g).value[:-1] != "0" or sheet_rez.cell(row=5, column=g).value[-1:] != "5":
            #     if round(int(sheet_rez.cell(column=g, row=5).value[-2:]) / 10) * 10 in range(1, 60):
            #         sheet_rez.cell(row=5, column=g).value = sheet_rez.cell(row=5, column=g).value[:-2] + str(
            #             round(int(sheet_rez.cell(column=g, row=5).value[-2:]) / 10) * 10)
            #     elif round(int(sheet_rez.cell(column=g, row=5).value[-2:]) / 10) * 10 in range(0, 6):
            #         sheet_rez.cell(row=5, column=g).value = sheet_rez.cell(row=5, column=g).value[:-2] + \
            #                                                 str(round(
            #                                                     int(sheet_rez.cell(column=g, row=5).value[-2:]) / 10) * 10) \
            #                                                 + "0"
            #     else:
            #         sheet_rez.cell(row=5, column=g).value = str(
            #             int(time_shift(sheet_rez.cell(row=5, column=g).value)[0]) + 1) + ":00"
        time_end = str(sheet_rez.cell(row=5, column=g).value)
        # print(time_end)
        # закрашиваем шифты - определяем первую и последнюю строки
        i_shift_start = nachalo_konec(sheet_rez, time_start, time_end)[0]
        i_shift_end = nachalo_konec(sheet_rez, time_start, time_end)[1]

        # закрашиваем смену
        for z in range(i_shift_start, i_shift_end + 1):
            sheet_rez.cell(row=z, column=g).fill = greenFill

        # расставляем перерывы
        # print(pereriv[time_per])

        len_per = 0
        if isinstance(pereriv[time_per], list):
            len_per = len(pereriv[time_per])
        elif time_per == 0:
            len_per = 0
        else:
            len_per = 1
        # for i_p in range(len_per):
        # print(len_per)
        #"""
        # проверяем файл фикс
        fix_count = 0
        for i_fix in range(1, sheet_fix.max_row + 1):
            if sheet_rez.cell(row=2, column=g).value == sheet_fix.cell(row=i_fix, column=4).value:
                if sheet_fix.cell(row=i_fix, column=2).value == "вместе" and sheet_fix.cell(row=i_fix, column=3).value == sheet_rez.cell(row=7, column=g).value:
                    #print(sheet_fix.cell(row=i_fix, column=3).value)
                    for n_fix in range(4, g):
                        if sheet_rez.cell(row=2, column=n_fix).value == sheet_fix.cell(row=i_fix, column=1).value:
                            for k_fix in range(10, sheet_rez.max_row + 1):
                                sheet_rez.cell(row=k_fix, column=g).value = sheet_rez.cell(row=k_fix,
                                                                                           column=n_fix).value
                                if sheet_rez.cell(row=k_fix, column=g).value == 5:
                                    sheet_rez.cell(row=k_fix, column=g).fill = yellowFill
                            fix_count = 1
                            break
                    break

            if sheet_rez.cell(row=2, column=g).value == sheet_fix.cell(row=i_fix, column=1).value:
                if sheet_fix.cell(row=i_fix, column=2).value == "вместе" and sheet_fix.cell(row=i_fix, column=3).value == sheet_rez.cell(row=7, column=g).value:
                    #print(sheet_fix.cell(row=i_fix, column=1).value)
                    for n_fix in range(4, g):
                        if sheet_rez.cell(row=2, column=n_fix).value == sheet_fix.cell(row=i_fix, column=3).value:
                            for k_fix in range(10, sheet_rez.max_row + 1):
                                sheet_rez.cell(row=k_fix, column=g).value = sheet_rez.cell(row=k_fix,
                                                                                           column=n_fix).value
                                if sheet_rez.cell(row=k_fix, column=g).value == 5:
                                    sheet_rez.cell(row=k_fix, column=g).fill = yellowFill
                            fix_count = 1
                    break
                elif sheet_fix.cell(row=i_fix, column=2).value == "фикс"and sheet_fix.cell(row=i_fix, column=3).value == sheet_rez.cell(row=7, column=g).value:
                    #print(sheet_fix.cell(row=i_fix, column=1).value)
                    for n_fix in range(4, sheet_fix.max_column + 1):
                        if sheet_fix.cell(row=i_fix, column=n_fix).value is not None:
                            start = shift(sheet_fix.cell(row=i_fix, column=n_fix).value)[0]
                            stop = shift(sheet_fix.cell(row=i_fix, column=n_fix).value)[1]
                            sheet_rez = shift_yellow(sheet_rez, g, start, stop)
                            #print(start, stop)
                    fix_count = 1
                    break
                elif sheet_fix.cell(row=i_fix, column=2).value == "последний"and sheet_fix.cell(row=i_fix, column=3).value == sheet_rez.cell(row=7, column=g).value:
                    #print(sheet_fix.cell(row=i_fix, column=1).value)
                    sheet_rez = posledniy(sheet_rez, len_per, time_per, i_shift_start, i_shift_end)
                            #print(start, stop)
                    fix_count = 1
                    break
                elif sheet_fix.cell(row=i_fix, column=2).value == "первый" and sheet_fix.cell(row=i_fix, column=3).value == sheet_rez.cell(row=7, column=g).value:
                    #print(sheet_fix.cell(row=i_fix, column=1).value)
                    sheet_rez = perviy(sheet_rez, len_per, time_per, i_shift_start, i_shift_end, sheet_fix, i_fix)
                            #print(start, stop)
                    fix_count = 1
                    break
                elif sheet_fix.cell(row=i_fix, column=2).value == "обед" and sheet_fix.cell(row=i_fix, column=3).value == sheet_rez.cell(row=7, column=g).value:
                    #print(sheet_fix.cell(row=i_fix, column=1).value)
                    # закрашиваем обед
                    start = shift(sheet_fix.cell(row=i_fix, column=4).value)[0]
                    stop = shift(sheet_fix.cell(row=i_fix, column=4).value)[1]
                    sheet_rez = shift_yellow(sheet_rez, g, start, stop)
                    # сравниваем начало смены и конец
                    shift_beg1 = nachalo_konec(sheet_rez, time_start, start)[0]
                    shift_end1 = nachalo_konec(sheet_rez, time_start, start)[1]
                    shift_beg2 = nachalo_konec(sheet_rez, stop, time_end)[0]
                    shift_end2 = nachalo_konec(sheet_rez, stop, time_end)[1]
                    #print(shift_end1 - shift_beg1,shift_end2 - shift_beg2)
                    # находим каким шифтом идет обед
                    for i_lunch in range(len_per):
                        if pereriv[time_per][i_lunch] == 6:
                            lunch_30 = i_lunch
                    #if (shift_end1 - shift_beg1) <= (shift_end2 - shift_beg2):
                    shift_per = (shift_end1 - shift_beg1 + shift_end2 - shift_beg2) / (len_per - 1)
                    #len_per1 = math.floor((shift_end1 - shift_beg1) / shift_per) if 0.25 <= ((shift_end1 - shift_beg1) / shift_per - math.floor((shift_end1 - shift_beg1) / shift_per)) < 0.75 else math.ceil((shift_end1 - shift_beg1) / shift_per)
                    len_per1 = round((shift_end1 - shift_beg1) / shift_per + 0.25)
                    len_per2 = len_per - 1 - len_per1
                    #print(len_per1, len_per2)
                    if len_per1 < 0:
                        len_per1 = 0
                        len_per2 -= 1
                    if len_per2 < 0:
                        len_per2 = 0
                        len_per1 -= 1
                    sheet_rez,  i_lunch_next = set_pereriv(sheet_rez, len_per1, time_per, shift_beg1, shift_end1, 0,lunch_30)
                    sheet_rez = set_pereriv(sheet_rez, len_per2, time_per, shift_beg2, shift_end2, i_lunch_next, lunch_30)[0]



                    fix_count = 1
                    break


        if fix_count == 1:
            continue
        #"""
        sheet_rez = set_pereriv(sheet_rez, len_per, time_per, i_shift_start, i_shift_end)[0]
    try:
        wb_rez.save(f"перерывы_сборка.xlsx")
        input("Перерывы расставлены. Открываем файл перерывы_сборка.xlsx и проверяем. Нажмите ENTER для продолжения...")
    except OSError:
        input(
            "Невозможно сохранить данные. Закройте файл перерывы_сборка.xlsx "
            "и запустите программу заново. Нажмите ENTER для продолжения...")

except Exception as err:
    print(err)
    input(
        f"ОШИБКА!!! Проверьте в перерыв_сборка данные по оператору {sheet_rez.cell(row=2, column=g).value}. Нажмите ENTER для продолжения...")
