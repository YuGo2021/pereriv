import openpyxl
import pereriv_lib as per

    
wb_grafik_per = openpyxl.load_workbook(f"перерывы_сборка.xlsx")
for s in range(len(wb_grafik_per.sheetnames)):
    if wb_grafik_per.sheetnames[s] == 'перерывы':
        wb_grafik_per.active = s
        break


print("Готовим таблицы для рассылки...")
sheet_per = wb_grafik_per.active


sheet_sm = per.new_sheet(wb_grafik_per, "Смирновка")
sheet_vis = per.new_sheet(wb_grafik_per, "Высота")
sheet_nsk = per.new_sheet(wb_grafik_per, "НСК")
sheet_kir = per.new_sheet(wb_grafik_per, "Киров")
sheet_rost = per.new_sheet(wb_grafik_per, "Ростов")
sheet_nn = per.new_sheet(wb_grafik_per, "НиНо")

# формируем рассылку для Смирновки
per.get_grafik(sheet_per, sheet_sm, "Смирновка")

# формируем рассылку для Высоты
per.get_grafik(sheet_per, sheet_vis, "Высота")
   
# формируем рассылку для Кирова
per.get_grafik(sheet_per, sheet_kir, "Киров")

# формируем рассылку для Новосибирска
per.get_grafik(sheet_per, sheet_nsk, "НСК")

# формируем рассылку для Ростов
per.get_grafik(sheet_per, sheet_rost, "Ростов")
 
# формируем рассылку для Нижнего Новгорода
per.get_grafik(sheet_per, sheet_nn, "НиНо")

 
try:
    wb_grafik_per.save(f"перерывы_сборка.xlsx") 
    input("Таблицы для рассылки подготовлены. Открываем файл перерывы_сборка.xlsx и "
          "проверяем. Нажмите ENTER для продолжения...")
except OSError:
    input("Невозможно сохранить данные. Закройте файл перерывы_сборка.xlsx и "
          "запустите программу заново. Нажмите ENTER для продолжения...")
