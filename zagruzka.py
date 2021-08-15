import re
import datetime as dt
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side

def shift(shift_cell):
    time_regex = r"((([0]{1})?([1-9]{1})\:([0-9]{2}))|(([1-2]{1}[0-9]{1})\:([0-9]{2})))\s?-\s?((([0]{1})?([1-9]{1})\:([0-9]{2}))|((([1-2]{1})([0-9]{1}))\:([0-9]{2})))"
    shift_test = re.search(time_regex, shift_cell)
    shift_start = ""
    shift_end = ""
    if shift_test is not None:
        if shift_test.group(4) is not None: 
            shift_start = f"{shift_test.group(4)}:{shift_test.group(5)}"
        else:
            shift_start = shift_test.group(1)
        
        if shift_test.group(10) is not None: 
            shift_end = f"{shift_test.group(12)}:{shift_test.group(13)}"
        else:
            shift_end = shift_test.group(9)
        #print(shift_start)
        
    return (shift_start , shift_end, shift_test.group(5), shift_test.group(13))
    
def fio(fio_cell):
    # преобразование ФИО в Ф И.О.
    if fio_cell == None:
         return ""
    else:
        fio_regex = r"([А-ЯЁ]{1}[а-яё]+)\s{1}(\(([А-ЯЁ]{1}[а-яё]+)\)\s{1})?([А-ЯЁ]{1}[а-яё]+)(\s{1}([А-ЯЁ]{1}[а-яё]+))?"
        fio_test = re.search(fio_regex, fio_cell)
        fio_new = ""
        if fio_test is not None:
            #print(fio_test.group(0))
            #print(fio_test.group(1))
            #print(fio_test.group(2))
            #print(fio_test.group(3))
            #print(fio_test.group(4))
            
            if fio_test.group(6) is not None:
                fio_new = fio_test.group(1)+" "+fio_test.group(4)[0]+". "+fio_test.group(6)[0]+"."
            else:
                fio_new = fio_test.group(1)+" "+fio_test.group(4)[0]+"."
            #print(fio_new)
    return fio_new

def time_shift(cell):
    time_regex = r"(\d{1,2})[\-\:]{1}(\d{1,2})"
    time_test = re.search(time_regex, cell)
    time_hour = ""
    time_min = ""
    if time_test is not None:
        time_hour = time_test.group(1)
        time_min = time_test.group(2)
    return (time_hour, time_min)

month_txt = input("Введите текстом месяц для анализа: ")

# открываем графики и делаем активной лист с нцжным месяцем
print("Проверяем график Смирновки...")
wb_grafik1 = openpyxl.load_workbook("График работы 2021 Смирновка.xlsx")
for s in range(len(wb_grafik1.sheetnames)):
    if (wb_grafik1.sheetnames[s].lower()).find(month_txt.lower()) != -1:
        print("Выбран лист ", wb_grafik1.sheetnames[s])
        break
wb_grafik1.active = s
sheet_grafik1 = wb_grafik1.active

print("Проверяем график Высоты...")
wb_grafik2 = openpyxl.load_workbook("График работы 2021 Высота.xlsx")
for s in range(len(wb_grafik2.sheetnames)):
    if (wb_grafik2.sheetnames[s].lower()).find(month_txt.lower()) != -1:
        print("Выбран лист ", wb_grafik2.sheetnames[s])
        break
wb_grafik2.active = s
sheet_grafik2 = wb_grafik2.active

print("Проверяем график Кирова...")
wb_grafik3 = openpyxl.load_workbook("График работы 2021 Киров .xlsx")
for s in range(len(wb_grafik3.sheetnames)):
    if (wb_grafik3.sheetnames[s].lower()).find(month_txt.lower()) != -1:
        print("Выбран лист ", wb_grafik3.sheetnames[s])
        break
wb_grafik3.active = s
sheet_grafik3 = wb_grafik3.active

print("Проверяем график Новосибирска...")
wb_grafik4 = openpyxl.load_workbook("График работы 2021 Новосибирск.xlsx")
for s in range(len(wb_grafik4.sheetnames)):
    if (wb_grafik4.sheetnames[s].lower()).find(month_txt.lower()) != -1:
        print("Выбран лист ", wb_grafik4.sheetnames[s])
        break
wb_grafik4.active = s
sheet_grafik4 = wb_grafik4.active

wb_rez = openpyxl.Workbook()
wb_rez.create_sheet(title = 'смены', index = 0)
sheet_rez = wb_rez['смены']

sheet_rez['A1'] = "День недели"
sheet_rez['A2'] = "Дата"
sheet_rez['A3'] = "Количество смен по графику"

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
medium = Side(border_style="medium", color="000000")

#Заполняем шапку с датами и днями недели
for i in range(1, sheet_grafik1.max_row + 1):
    if sheet_grafik1.cell(row = i, column = 6).value == "ОПЕРАТОРЫ 5/2":
        for j in range(9,sheet_grafik1.max_column + 1):
            if  sheet_grafik1.cell(row = i, column = j).value is not None:            
                sheet_rez.cell(row = 1, column = j-4).value = sheet_grafik1.cell(row = i, column = j).value
                sheet_rez.cell(row = 2, column = j-4).value = sheet_grafik1.cell(row = i+1, column = j).value
            else:
                break
        break


#Заполняем таблицу с выходами по сменам по датам месяца
i_5_2 = 0
i_2_2 = 0
i_1_2 = 0

# собираем Смирновку
for i in range(1, sheet_grafik1.max_row + 1):
    if sheet_grafik1.cell(row = i, column = 6).value == "ОПЕРАТОРЫ 5/2":
        i_5_2 = i

i_op = 4 # счетсчик заполнения строк

for k in range(i_5_2,  sheet_grafik1.max_row + 1): 
    if sheet_grafik1.cell(row = k, column = 4).value is not None and sheet_grafik1.cell(row = k, column = 6).value is not None and sheet_grafik1.cell(row = k, column = 5).value is not None:
        sheet_rez.cell(row = i_op, column = 1).value = "Смирновка"
        sheet_rez.cell(row = i_op, column = 2).value = shift(sheet_grafik1.cell(row = k, column = 5).value)[0]
        sheet_rez.cell(row = i_op, column = 3).value = shift(sheet_grafik1.cell(row = k, column = 5).value)[1]
        sheet_rez.cell(row = i_op, column = 4).value = fio(sheet_grafik1.cell(row = k, column = 6).value)
        print(sheet_rez.cell(row = i_op, column = 4).value)
        for sm_st in range(5,sheet_rez.max_column + 1):
            if sheet_grafik1.cell(row = k-1, column = sm_st+4).value == "С" or sheet_grafik1.cell(row = k-1, column = sm_st+4).value == "C" :
                sheet_rez.cell(row = i_op, column = sm_st).value = 1
        i_op += 1

# сибираем Высоту
for i in range(1, sheet_grafik2.max_row + 1):
    if sheet_grafik2.cell(row = i, column = 6).value == "ОПЕРАТОРЫ 5/2":
        i_5_2 = i


for k in range(i_5_2,  sheet_grafik2.max_row + 1): 
    if sheet_grafik2.cell(row = k, column = 4).value is not None and sheet_grafik2.cell(row = k, column = 6).value is not None and sheet_grafik2.cell(row = k, column = 5).value is not None:
        sheet_rez.cell(row = i_op, column = 1).value = "Высота"
        sheet_rez.cell(row = i_op, column = 2).value = shift(sheet_grafik2.cell(row = k, column = 5).value)[0]
        sheet_rez.cell(row = i_op, column = 3).value = shift(sheet_grafik2.cell(row = k, column = 5).value)[1]
        sheet_rez.cell(row = i_op, column = 4).value = fio(sheet_grafik2.cell(row = k, column = 6).value)
        print(sheet_rez.cell(row = i_op, column = 4).value)
        for sm_st in range(5,sheet_rez.max_column + 1):
            if sheet_grafik2.cell(row = k-1, column = sm_st+4).value == "С" or sheet_grafik2.cell(row = k-1, column = sm_st+4).value == "C" :
                sheet_rez.cell(row = i_op, column = sm_st).value = 1
        i_op += 1
        
# сибираем Киров
for i in range(1, sheet_grafik3.max_row + 1):
    if sheet_grafik3.cell(row = i, column = 6).value == "ОПЕРАТОРЫ 2/2":
        i_5_2 = i


for k in range(i_5_2,  sheet_grafik3.max_row + 1): 
    if sheet_grafik3.cell(row = k, column = 4).value is not None and sheet_grafik3.cell(row = k, column = 6).value is not None and sheet_grafik3.cell(row = k, column = 5).value is not None:
        sheet_rez.cell(row = i_op, column = 1).value = "Киров"
        sheet_rez.cell(row = i_op, column = 2).value = shift(sheet_grafik3.cell(row = k, column = 5).value)[0]
        sheet_rez.cell(row = i_op, column = 3).value = shift(sheet_grafik3.cell(row = k, column = 5).value)[1]
        sheet_rez.cell(row = i_op, column = 4).value = fio(sheet_grafik3.cell(row = k, column = 6).value)
        print(sheet_rez.cell(row = i_op, column = 4).value)
        for sm_st in range(5,sheet_rez.max_column + 1):
            if sheet_grafik3.cell(row = k-1, column = sm_st+4).value == "С" or sheet_grafik3.cell(row = k-1, column = sm_st+4).value == "C" :
                sheet_rez.cell(row = i_op, column = sm_st).value = 1
        i_op += 1

# сибираем Новосибирск
for i in range(1, sheet_grafik4.max_row + 1):
    if sheet_grafik4.cell(row = i, column = 6).value == "ОПЕРАТОРЫ 2/2":
        i_5_2 = i


for k in range(i_5_2,  sheet_grafik4.max_row + 1): 
    if sheet_grafik4.cell(row = k, column = 4).value is not None and sheet_grafik4.cell(row = k, column = 6).value is not None and sheet_grafik4.cell(row = k, column = 5).value is not None:
        sheet_rez.cell(row = i_op, column = 1).value = "НСК"
        sheet_rez.cell(row = i_op, column = 2).value = f"{int(time_shift(shift(sheet_grafik4.cell(row = k, column = 5).value)[0])[0])-4}:{time_shift(shift(sheet_grafik4.cell(row = k, column = 5).value)[0])[1]}"
        sheet_rez.cell(row = i_op, column = 3).value = f"{int(time_shift(shift(sheet_grafik4.cell(row = k, column = 5).value)[1])[0])-4}:{time_shift(shift(sheet_grafik4.cell(row = k, column = 5).value)[1])[1]}"
        sheet_rez.cell(row = i_op, column = 4).value = fio(sheet_grafik4.cell(row = k, column = 6).value)
        print(sheet_rez.cell(row = i_op, column = 4).value)
        for sm_st in range(5,sheet_rez.max_column + 1):
            if sheet_grafik4.cell(row = k-1, column = sm_st+4).value == "С" or sheet_grafik4.cell(row = k-1, column = sm_st+4).value == "C" :
                sheet_rez.cell(row = i_op, column = sm_st).value = 1
        i_op += 1


for i in range(5, sheet_rez.max_column +1):
    sheet_rez.cell(row = 3, column = i).value = f"=SUM({get_column_letter(i)}4:{get_column_letter(i)}{sheet_rez.max_row +1})"
    
    
for i in range(1, sheet_rez.max_column +1):
    for j in range(1, sheet_rez.max_row + 1):
        sheet_rez.cell(row = j, column = i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
if wb_rez.sheetnames.count('Получасовая') == 0:
    wb_rez.create_sheet(title = 'Получасовая', index = 0)
sheet_ch = wb_rez['Получасовая']

for row in sheet_ch['A1:AF52']:
    for cell in row:
            cell.value = None
    
sheet_ch['A1'] = month_txt
# заполняем верхнюю таблицу
for i in range(5, sheet_rez.max_column+1):
    sheet_ch.cell(row = 1, column = i-3).value = sheet_rez.cell(row = 1, column = i).value
    sheet_ch.cell(row = 2, column = i-3).value = sheet_rez.cell(row = 2, column = i).value
#заполняем время  в первых столбцах
for i in range(3, 50, 2):
    sheet_ch.cell(row = i, column = 1).value = f"{(i-3)//2}:00"
    sheet_ch.cell(row = i+1, column = 1).value = f"{(i-3)//2}:30"


# начинаем сборку
for i in range(4,sheet_rez.max_row+1):
    for l in range(3,51):
        if sheet_rez.cell(row = i, column = 2).value == sheet_ch.cell(row = l, column = 1).value:
            time_start_i = l
        if sheet_rez.cell(row = i, column = 3).value == sheet_ch.cell(row = l, column = 1).value:
            time_end_i = l
    for r_i in range(time_start_i, time_end_i):
        for c_i in range(2,sheet_ch.max_column+1):
            if sheet_rez.cell(row = i, column = c_i+3).value == 1:
                if sheet_ch.cell(row = r_i, column = c_i).value is not None: 
                    sheet_ch.cell(row = r_i, column = c_i).value = int(sheet_ch.cell(row = r_i, column = c_i).value) + 1
                else:
                    sheet_ch.cell(row = r_i, column = c_i).value = 1

                    
# делаем почасовую разбивку                    
if wb_rez.sheetnames.count('Почасовая') == 0:
    wb_rez.create_sheet(title = 'Почасовая', index = 0)
sheet_ch1 = wb_rez['Почасовая']

for row in sheet_ch1['A1:AF52']:
    for cell in row:
            cell.value = None
    
sheet_ch1['A1'] = month_txt
# заполняем верхнюю таблицу
for i in range(5, sheet_rez.max_column+1):
    sheet_ch1.cell(row = 1, column = i-3).value = sheet_rez.cell(row = 1, column = i).value
    sheet_ch1.cell(row = 2, column = i-3).value = sheet_rez.cell(row = 2, column = i).value
#заполняем время  в первых столбцах
for i in range(3, 27):
    sheet_ch1.cell(row = i, column = 1).value = f"{(i-3)}:00"

for i_r in range(3,51,2):
    for i_c in range(2,sheet_ch.max_column+1):
        if sheet_ch.cell(row = i_r, column = i_c).value is not None and sheet_ch.cell(row = i_r+1, column = i_c).value is not None:
            sheet_ch1.cell(row = i_r//2+2, column = i_c).value = (sheet_ch.cell(row = i_r, column = i_c).value + sheet_ch.cell(row = i_r+1, column = i_c).value)//2
    
try:        
    wb_rez.save(f"План.xlsx")        
    input("Графики собраны. Проверяем...")
except OSError:
    input("Невозможно сохранить данные. Закройте файл перерывы_сборка.xlsx и запустите программу заново. Нажмите ENTER для продолжения...")    