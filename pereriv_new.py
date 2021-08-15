import re
import datetime as dt
import time
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side

def time_shift(cell):
    time_regex = r"(\d{1,2})[\-\:]{1}(\d{1,2})"
    time_test = re.search(time_regex, cell)
    time_hour = ""
    time_min = ""
    if time_test is not None:
        time_hour = time_test.group(1)
        time_min = time_test.group(2)
    return (time_hour, time_min)

def shift(shift_cell):
    time_regex = r"((([0]{1})?([1-9]{1})\:([0-9]{2}))|(([1-2]{1}[0-9]{1})\:([0-9]{2})))\s?-\s?((([0]{1})?([1-9]{1})\:([0-9]{2}))|((([1-2]{1})([0-9]{1}))\:([0-9]{2})))"
    shift_test = re.search(time_regex, shift_cell)
    shift_start = ""
    shift_end = ""
    if shift_test is not None:
        if shift_test.group(4) is not None: 
            shift_start = f"{shift_test.group(4)}:{shift_test.group(5)}"
        else:
            shift_start = shift_test.group(1)
        
        if shift_test.group(10) is not None: 
            shift_end = f"{shift_test.group(12)}:{shift_test.group(13)}"
        else:
            shift_end = shift_test.group(9)
        #print(shift_start)
        
    return (shift_start , shift_end, shift_test.group(5), shift_test.group(13))
        

wb_grafik_per = openpyxl.load_workbook(f"перерывы_сборка.xlsx")
for s in range(len(wb_grafik_per.sheetnames)):
    if wb_grafik_per.sheetnames[s] == 'перерывы':
        break
wb_grafik_per.active = s

sheet_per = wb_grafik_per.active

wb_graf = openpyxl.Workbook()
wb_graf.create_sheet(title = 'график', index = 0)
sheet_graf = wb_graf['график']
sheet_graf.delete_rows(1, 200)

# заполняем первые 3 строки 
sheet_graf.merge_cells("A1:A3")
sheet_graf["A1"] = "Площадка"
sheet_graf.merge_cells("B1:B3")
sheet_graf["B1"] = "Время"
sheet_graf.merge_cells("C1:C3")
sheet_graf["C1"] = "ФИО оператора"

count_hour = 0        
for m in range (4, 100, 4):
    sheet_graf.merge_cells(f"{get_column_letter(m)}1:{get_column_letter(m+3)}1")
    sheet_graf[f"{get_column_letter(m)}1"] = f"=SUM({get_column_letter(m)}4:{get_column_letter(m+3)}200)/4"
    sheet_graf.merge_cells(f"{get_column_letter(m)}2:{get_column_letter(m+3)}2")
    sheet_graf[f"{get_column_letter(m)}2"] = count_hour
    count_hour += 1
    sheet_graf[f"{get_column_letter(m)}3"] = 0
    sheet_graf[f"{get_column_letter(m+1)}3"] = 15
    sheet_graf[f"{get_column_letter(m+2)}3"] = 30
    sheet_graf[f"{get_column_letter(m+3)}3"] = 45

# копируем данные из зеленой таблицы с перерывами
for i in range(4, sheet_per.max_column + 1): 
        sheet_graf[f"A{i}"] = sheet_per[f"{get_column_letter(i)}1"].value
        sheet_graf[f"B{i}"] = "%s%s%s" % (sheet_per[f"{get_column_letter(i)}4"].value , "-", sheet_per[f"{get_column_letter(i)}5"].value)
        sheet_graf[f"C{i}"] = sheet_per[f"{get_column_letter(i)}2"].value
        sm_beg = 4
        for j in range(4,sheet_graf.max_column + 1,4):
            if time_shift(str(sheet_per[f"{get_column_letter(i)}4"].value))[0] == str(sheet_graf[f"{get_column_letter(j)}2"].value):
                if time_shift(str(sheet_per[f"{get_column_letter(i)}4"].value))[1] == "00":
                    sm_beg = j
                    break
                elif time_shift(str(sheet_per[f"{get_column_letter(i)}4"].value))[1] == "30":
                    sm_beg = j+2
                    break
        
        for z in range(sm_beg,sheet_graf.max_column + 1):
            sheet_graf[f"{get_column_letter(z)}{i}"] = 1
            if time_shift(str(sheet_per[f"{get_column_letter(i)}5"].value))[0] == str(sheet_graf[f"{get_column_letter(z)}2"].value):
                if time_shift(str(sheet_per[f"{get_column_letter(i)}5"].value))[1] == "00":
                    sheet_graf[f"{get_column_letter(z)}{i}"] = None
                    break
                elif time_shift(str(sheet_per[f"{get_column_letter(i)}5"].value))[1] == "30":
                    sheet_graf[f"{get_column_letter(z+1)}{i}"] = 1
                    break
                elif time_shift(str(sheet_per[f"{get_column_letter(i)}5"].value))[1] == "25":
                    sheet_graf[f"{get_column_letter(z+1)}{i}"] = 2/3
                    break                    
        # зполняем перерывы
        for strok in range(10, sheet_per.max_row + 1): 
            if sheet_per.cell(row = strok, column = i).value == 5:
                for strok1 in range(strok, strok - 12, -1):
                    if sheet_per.cell(row = strok1, column = 1).value is not None:
                        hour = time_shift(shift(sheet_per.cell(row = strok1, column = 1).value)[0])[0]
                min = time_shift(sheet_per.cell(row = strok, column = 2).value)[0]
                for ii in range(4,sheet_graf.max_column + 1,4):
                    if hour == str(sheet_graf[f"{get_column_letter(ii)}2"].value):
                        for jj in range(ii,ii+3):
                            if  int(sheet_graf[f"{get_column_letter(jj)}3"].value) <= int(min) < int(sheet_graf[f"{get_column_letter(jj+1)}3"].value):
                                sheet_graf[f"{get_column_letter(jj)}{i}"].value -= 1/3
                        if  45 <= int(min) < 60:
                            sheet_graf[f"{get_column_letter(ii+3)}{i}"].value -= 1/3
    
# рисуем границы
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
medium = Side(border_style="medium", color="000000")
for i in range(1, sheet_graf.max_column + 1):
    for j in range(1, sheet_graf.max_row + 1):
        sheet_graf.cell(row = j, column = i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    

try:            
    wb_graf.save(f"график.xlsx")        
    input("Графики собраны. Проверяем...")    
except OSError:
    input("Невозможно сохранить данные. Закройте файл график.xlsx и запустите программу заново. Нажмите ENTER для продолжения...")
