import re
import datetime as dt
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side


def fio(fio_cell):
    # преобразование ФИО в Ф И.О.
    if fio_cell is None:
        return ""
    else:
        fio_regex = r"([А-ЯЁ]{1}[а-яё]+)\s{1}(\(([А-ЯЁ]{1}[а-яё]+)\)\s{1})?([А-ЯЁ]{1}[а-яё]+)(\s{1}([А-ЯЁ]{1}[а-яё]+))?"
        fio_test = re.search(fio_regex, fio_cell)
        fio_new = ""
        if fio_test is not None:
            # print(fio_test.group(0))
            # print(fio_test.group(1))
            # print(fio_test.group(2))
            # print(fio_test.group(3))
            # print(fio_test.group(4))

            if fio_test.group(6) is not None:
                fio_new = fio_test.group(1) + " " + fio_test.group(4)[0] + ". " + fio_test.group(6)[0] + "."
            else:
                fio_new = fio_test.group(1) + " " + fio_test.group(4)[0] + "."
            # print(fio_new)
    return fio_new


def oper_div(sheet_name):  # поиск в столбце с ФИО разделителей по сменам
    for i in range(1, sheet_name.max_row + 1):
        if sheet_name.cell(row=i, column=6).value == "ОПЕРАТОРЫ 5/2":
            i_5_2 = i
        # print(i_5_2)
        if sheet_name.cell(row=i, column=6).value == "ОПЕРАТОРЫ 2/2":
            i_2_2 = i
        # print(i_2_2)
        if (sheet_name.cell(row=i, column=6).value == "НЕПОЛНЫЙ РАБОЧИЙ ДЕНЬ" or
            sheet_name.cell(row=i, column=6).value == "ОПЕРАТОРЫ  неполного рабочего  дня"):
            i_1_2 = i
        # print(i_1_2)
    return i_5_2, i_2_2, i_1_2


def fio_full(fio_cell):
    # преобразование ФИО в Ф И.О.
    if fio_cell is None:
        return ""
    else:
        fio_regex = r"([А-ЯЁ]{1}[а-яё]+)\s{1}(\(([А-ЯЁ]{1}[а-яё]+)\)\s{1})?([А-ЯЁ]{1}[а-яё]+)(\s{1}([А-ЯЁ]{1}[а-яё]+))?"
        fio_test = re.search(fio_regex, fio_cell)
        fio_new = ""
        if fio_test is not None:
            # print(fio_test.group(0))
            # print(fio_test.group(1))
            # print(fio_test.group(2))
            # print(fio_test.group(3))
            # print(fio_test.group(4))

            if fio_test.group(6) is not None:
                fio_new = fio_test.group(0)
            else:
                fio_new = fio_test.group(1) + " " + fio_test.group(4)
            # print(fio_
    return fio_new


def shift(shift_cell):
    time_regex = r"((([0]{1})?([1-9]{1})\:([0-9]{2}))|(([1-2]{1}[0-9]{1})\:" \
                 r"([0-9]{2})))\s?-\s?((([0]{1})?([1-9]{1})\:([0-9]{2}))|((([1-2]{1})([0-9]{1}))\:([0-9]{2})))"
    shift_test = re.search(time_regex, shift_cell)
    shift_start = ""
    shift_end = ""
    if shift_test is not None:
        if shift_test.group(4) is not None:
            shift_start = f"{shift_test.group(4)}:{shift_test.group(5)}"
        else:
            shift_start = shift_test.group(1)

        if shift_test.group(10) is not None:
            shift_end = f"{shift_test.group(12)}:{shift_test.group(13)}"
        else:
            shift_end = shift_test.group(9)
        # print(shift_start)

    return shift_start, shift_end, shift_test.group(5), shift_test.group(13)


def time_shift(cell):
    time_regex = r"(\d{1,2})[\-\:]{1}(\d{1,2})"
    time_test = re.search(time_regex, cell)
    time_hour = ""
    time_min = ""
    if time_test is not None:
        time_hour = time_test.group(1)
        time_min = time_test.group(2)
    return time_hour, time_min


def nsk(cell):
    time_regex = r"(\d{1,2})[\-\:]{1}(\d{1,2})"
    time_test = re.search(time_regex, cell)
    time_hour = ""
    time_min = ""
    if time_test is not None:
        time_hour = str(int(time_test.group(1)) - 4)
        time_min = time_test.group(2)
    return f"{time_hour}:{time_min}"


def sting_no(sheet):
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=6).value == "ОПЕРАТОРЫ 5/2":
            i_5_2 = i
            # print(i_5_2)
        if sheet.cell(row=i, column=6).value == "ОПЕРАТОРЫ 2/2":
            i_2_2 = i
            # print(i_2_2)
        if (sheet.cell(row=i, column=6).value == "НЕПОЛНЫЙ РАБОЧИЙ ДЕНЬ" or
                sheet.cell(row=i, column=6).value == "ОПЕРАТОРЫ  неполного рабочего  дня"):
            i_1_2 = i
            # print(i_1_2)
    return (i_5_2, i_2_2, i_1_2)


def find_cell(k, sheet, CC_name, shift):
    wb_per = openpyxl.load_workbook("Исключения.xlsx")
    sheet_per = wb_per.active

    # n1 = ""
    # n2 = ""
    # n3 = ""
    # n4 = ""
    # n5 = ""
    n6 = 0
    # n7 = 0
    color = 0
    # oper = ""
    oper = fio_full(sheet.cell(row=k, column=6).value)
    # print(oper)
    n1 = CC_name
    n2 = fio(oper)
    n3 = shift
    # print(type(sheet.cell(row = k, column = 5).value))
    # temp = sheet.cell(row = k, column = 5).value
    # print(shift(temp)[0])
    # n4 = shift(sheet.cell(row = k, column = 5).value)[0]
    # n5 = shift(sheet.cell(row = k, column = 5).value)[1]
    cell = sheet.cell(row=k, column=9).value
    cell_plus_1 = sheet.cell(row=k + 1, column=9).value

    if (type(cell) == int or type(cell) == float) and (type(cell_plus_1) == int or type(cell_plus_1) == float):
        n4 = "с+доп"
    elif (type(cell) == int or type(cell) == float) and not (type(cell_plus_1) == int or type(cell_plus_1) == float):
        n4 = "c"
    elif (type(cell_plus_1) == int or type(cell_plus_1) == float) and not (type(cell) == int or type(cell) == float):
        n4 = "доп"

    if (type(cell) == int or type(cell) == float) and (type(cell_plus_1) == int or type(cell_plus_1) == float):
        n6 = sheet.cell(row=k, column=9).value + sheet.cell(row=k + 1, column=9).value
    elif (type(cell) == int or type(cell) == float) and not (type(cell_plus_1) == int or type(cell_plus_1) == float):
        n6 = sheet.cell(row=k, column=9).value
    elif (type(cell_plus_1) == int or type(cell_plus_1) == float) and not (type(cell) == int or type(cell) == float):
        n6 = sheet.cell(row=k + 1, column=9).value

    #     if sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is not None:
    #         n4 = "с+доп"
    #     elif sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is None:
    #         n4 = "c"
    #     else:
    #         n4 = "доп"

    #     if sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is not None:
    #         n6 = sheet.cell(row = k, column = 9).value + sheet.cell(row = k+1, column = 9).value
    #     elif sheet.cell(row = k, column = 9).value is not None and sheet.cell(row = k+1, column = 9).value is None:
    #         n6 = sheet.cell(row = k, column = 9).value
    #     else:
    #         n6 = sheet.cell(row = k+1, column = 9).value

    if n4 == "c":
        if n6 == 11:
            n7 = 75

        elif n6 > 11:
            n7 = 75
            color = 1
        elif n6 == 8:
            n7 = 65
        elif 8 < n6 < 11:
            n7 = 70
            color = 1
        elif n6 == 3:
            n7 = 15
            color = 1
        elif 3 < n6 < 6:
            n7 = 20
            color = 1
        elif n6 == 6:
            n7 = 30
            color = 1
        elif 6 < n6 < 8:
            n7 = 60
            color = 1
        else:
            n7 = ""
            color = 1

    else:
        if n6 == 11:
            n7 = 75
        elif 8 < n6 < 11:
            n7 = 70

        elif n6 > 11:
            n7 = 75

        elif n6 == 8:
            n7 = 65
        elif 3 < n6 < 6:
            n7 = 20

        elif n6 == 3:
            n7 = 15

        elif n6 == 6:
            n7 = 30

        elif 6 < n6 < 8:
            n7 = 60

        else:
            n7 = ""
        color = 1
    for i in range(2, sheet_per.max_row + 1):
        if n2 == sheet_per.cell(row=i, column=1).value:
            n7 = sheet_per.cell(row=i, column=2).value
            if n4 == "c":
                color = 0
            else:
                color = 1
    return n1, n2, n3, n4, n6, n7, oper, color


def sum_cell(line1, line2, sheet):
    sum_lines = 0
    for z in range(4, sheet.max_column + 1):
        for line in range(line1, line2):
            if sheet.cell(row=line, column=z).value is not None:
                sum_lines += sheet.cell(row=line, column=z).value
    return sum_lines


def pereriv_CC(sheet_grafik, i_op1, sheet_rez1, CC_name):
    redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')

    i_5_2 = 0
    i_2_2 = 0
    i_1_2 = 0
    for i in range(1, sheet_grafik.max_row + 1):
        if sheet_grafik.cell(row=i, column=6).value == "ОПЕРАТОРЫ 5/2":
            i_5_2 = i
            # print(i_5_2)
        if sheet_grafik.cell(row=i, column=6).value == "ОПЕРАТОРЫ 2/2":
            i_2_2 = i
            # print(i_2_2)
        if (sheet_grafik.cell(row=i, column=6).value == "НЕПОЛНЫЙ РАБОЧИЙ ДЕНЬ" or
                sheet_grafik.cell(row=i, column=6).value == "ОПЕРАТОРЫ  неполного рабочего  дня"):
            i_1_2 = i
            # print(i_1_2)

    if i_5_2 > 0:
        if i_2_2 == 0:
            if i_1_2 == 0:
                i_1_2 = sheet_grafik.max_row + 4
    if i_2_2 == 0:
        if i_1_2 == 0:
            i_1_2 = sheet_grafik.max_row + 4
    if i_1_2 == 0:
        i_1_2 = sheet_grafik.max_row + 4

    # oper = ""
    if i_op1 < 4:
        i_op1 = 4


    if i_5_2 > 0:
        for k in range(i_5_2 + 1, i_2_2 - 3):
            if sheet_grafik.cell(row=k, column=4).value is not None \
                    and sheet_grafik.cell(row=k, column=6).value is not None \
                    and sheet_grafik.cell(row=k, column=5).value is not None \
                    and (type(sheet_grafik.cell(row=k, column=9).value) == int or type(
                sheet_grafik.cell(row=k, column=9).value) == float or type(
                sheet_grafik.cell(row=k + 1, column=9).value) == int or type(
                sheet_grafik.cell(row=k + 1, column=9).value) == float):
                oper = fio_full(sheet_grafik.cell(row=k, column=6).value)
                print(oper)
                sheet_rez1.cell(row=1, column=i_op1).value = CC_name
                sheet_rez1.cell(row=2, column=i_op1).value = fio(oper)
                sheet_rez1.cell(row=3, column=i_op1).value = "5/2"
                if CC_name == "НСК":
                    sheet_rez1.cell(row=4, column=i_op1).value = nsk(shift(sheet_grafik.cell(row=k, column=5).value)[0])
                # print(shift(sheet_grafik.cell(row = k, column = 5).value)[0])
                    sheet_rez1.cell(row=5, column=i_op1).value = nsk(shift(sheet_grafik.cell(row=k, column=5).value)[1])
                else:
                    sheet_rez1.cell(row=4, column=i_op1).value = shift(sheet_grafik.cell(row=k, column=5).value)[0]
                    # print(shift(sheet_grafik.cell(row = k, column = 5).value)[0])
                    sheet_rez1.cell(row=5, column=i_op1).value = shift(sheet_grafik.cell(row=k, column=5).value)[1]

                # print(shift(sheet_grafik.cell(row = k, column = 5).value)[1])
                sheet_rez1.cell(row=6, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "5/2")[3]
                sheet_rez1.cell(row=7, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "5/2")[4]
                sheet_rez1.cell(row=8, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "5/2")[5]
                if find_cell(k, sheet_grafik, CC_name, "5/2")[7] == 1:
                    sheet_rez1.cell(row=8, column=i_op1).fill = redFill
                sheet_rez1.cell(row=9,
                                column=i_op1).value = f"=SUM({get_column_letter(i_op1)}10:{get_column_letter(i_op1)}273)"
                # if sheet_grafik.cell(row = k, column = 9).value is not None and sheet_grafik.cell(row = k+1, column = 9).value is not None:
                #   sheet_rez1.cell(row = 6, column = i_op1).value = sheet_grafik.cell(row = k, column = 9).value + sheet_grafik.cell(row = k+1, column = 9).value
                # elif sheet_grafik.cell(row = k, column = 9).value is not None and sheet_grafik.cell(row = k+1, column = 9).value is None:
                #   sheet_rez1.cell(row = 6, column = i_op1).value = sheet_grafik.cell(row = k, column = 9).value
                # else:
                #   sheet_rez1.cell(row = 6, column = i_op1).value = sheet_grafik.cell(row = k+1, column = 9).value

                i_op1 += 1
    if i_2_2 > 0:

        for k in range(i_2_2 + 1, i_1_2 - 3):
            if sheet_grafik.cell(row=k, column=4).value is not None \
                    and sheet_grafik.cell(row=k, column=6).value is not None \
                    and sheet_grafik.cell(row=k, column=5).value is not None \
                    and (type(sheet_grafik.cell(row=k, column=9).value) == int or type(
                    sheet_grafik.cell(row=k, column=9).value) == float or type(
                    sheet_grafik.cell(row=k + 1, column=9).value) == int or type(
                    sheet_grafik.cell(row=k + 1, column=9).value) == float):
                oper = fio_full(sheet_grafik.cell(row=k, column=6).value)
                print(oper)
                sheet_rez1.cell(row=1, column=i_op1).value = CC_name
                sheet_rez1.cell(row=2, column=i_op1).value = fio(oper)
                sheet_rez1.cell(row=3, column=i_op1).value = "2/2"
                if CC_name == "НСК":
                    sheet_rez1.cell(row=4, column=i_op1).value = nsk(shift(sheet_grafik.cell(row=k, column=5).value)[0])
                    # print(shift(sheet_grafik.cell(row = k, column = 5).value)[0])
                    sheet_rez1.cell(row=5, column=i_op1).value = nsk(shift(sheet_grafik.cell(row=k, column=5).value)[1])
                else:
                    sheet_rez1.cell(row=4, column=i_op1).value = shift(sheet_grafik.cell(row=k, column=5).value)[0]
                    # print(shift(sheet_grafik.cell(row = k, column = 5).value)[0])
                    sheet_rez1.cell(row=5, column=i_op1).value = shift(sheet_grafik.cell(row=k, column=5).value)[1]
                sheet_rez1.cell(row=6, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "2/2")[3]
                sheet_rez1.cell(row=7, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "2/2")[4]
                sheet_rez1.cell(row=8, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "2/2")[5]
                if find_cell(k, sheet_grafik, CC_name, "5/2")[7] == 1:
                    sheet_rez1.cell(row=8, column=i_op1).fill = redFill
                sheet_rez1.cell(row=9,
                                column=i_op1).value = f"=SUM({get_column_letter(i_op1)}10:{get_column_letter(i_op1)}273)"
                i_op1 += 1
    if i_1_2 > 0:
        for k in range(i_1_2 + 1, sheet_grafik.max_row + 1):
            if sheet_grafik.cell(row=k, column=4).value is not None \
                    and sheet_grafik.cell(row=k, column=6).value is not None \
                    and sheet_grafik.cell(row=k, column=5).value is not None \
                    and (type(sheet_grafik.cell(row=k, column=9).value) == int or type(
                sheet_grafik.cell(row=k, column=9).value) == float or type(
                sheet_grafik.cell(row=k + 1, column=9).value) == int or type(
                sheet_grafik.cell(row=k + 1, column=9).value) == float):
                oper = fio_full(sheet_grafik.cell(row=k, column=6).value)
                print(oper)
                sheet_rez1.cell(row=1, column=i_op1).value = CC_name
                sheet_rez1.cell(row=2, column=i_op1).value = fio(oper)
                sheet_rez1.cell(row=3, column=i_op1).value = "1/2"
                if CC_name == "НСК":
                    sheet_rez1.cell(row=4, column=i_op1).value = nsk(shift(sheet_grafik.cell(row=k, column=5).value)[0])
                    # print(shift(sheet_grafik.cell(row = k, column = 5).value)[0])
                    sheet_rez1.cell(row=5, column=i_op1).value = nsk(shift(sheet_grafik.cell(row=k, column=5).value)[1])
                else:
                    sheet_rez1.cell(row=4, column=i_op1).value = shift(sheet_grafik.cell(row=k, column=5).value)[0]
                    # print(shift(sheet_grafik.cell(row = k, column = 5).value)[0])
                    sheet_rez1.cell(row=5, column=i_op1).value = shift(sheet_grafik.cell(row=k, column=5).value)[1]
                sheet_rez1.cell(row=6, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "1/2")[3]
                sheet_rez1.cell(row=7, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "1/2")[4]
                sheet_rez1.cell(row=8, column=i_op1).value = find_cell(k, sheet_grafik, CC_name, "1/2")[5]
                if find_cell(k, sheet_grafik, CC_name, "5/2")[7] == 1:
                    sheet_rez1.cell(row=8, column=i_op1).fill = redFill
                sheet_rez1.cell(row=9,
                                column=i_op1).value = f"=SUM({get_column_letter(i_op1)}10:{get_column_letter(i_op1)}273)"
                i_op1 += 1
    return sheet_rez1, i_op1


def new_sheet(wb_grafik, CC_name):
    if wb_grafik.sheetnames.count(CC_name) == 0:
        wb_grafik.create_sheet(title=CC_name, index=0)
    sheet = wb_grafik[CC_name]
    sheet['A1'] = "ФИО"
    sheet['B1'] = "Время работы"
    sheet['C1'] = "Перерыв 1"
    sheet['D1'] = "Перерыв 2"
    sheet['E1'] = "Перерыв 3"
    sheet['F1'] = "Перерыв 4"
    sheet['G1'] = "Перерыв 5"
    return sheet

def new_sheet_oktel(wb_grafik, CC_name):
    if wb_grafik.sheetnames.count(CC_name) == 0:
        wb_grafik.create_sheet(title=CC_name, index=0)
    sheet = wb_grafik[CC_name]
    sheet['A1'] = "name"
    sheet['B1'] = "wtime"
    sheet['C1'] = "break1"
    sheet['D1'] = "break2"
    sheet['E1'] = "break3"
    sheet['F1'] = "break4"
    sheet['G1'] = "break5"
    return sheet


def time_chek(cell):
    # print(type(cell))
    if type(cell) == dt.time:
        time_start = cell.strftime("%H:%M")
        # print(time_start)
        if time_start[0] == "0":
            time_start = time_start[1:]
    else:
        time_start = cell
        if time_start[5:7] == ":00":
            time_start = time_start[0:4]
        if time_start[0] == "0":
            time_start = time_start[1:]
    return time_start


def get_grafik(sheet_per, sheet_list, CC_name):
    # формируем рассылку для CC_name
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    medium = Side(border_style="medium", color="000000")
    no = 2
    for kol in range(4, sheet_per.max_column + 1):

        if sheet_per.cell(row=1, column=kol).value == CC_name:
            sheet_list.cell(row=no, column=1).value = sheet_per.cell(row=2, column=kol).value
            if CC_name == "НСК":
                time_start = f"{int(time_chek(sheet_per.cell(row=4, column=kol).value)[:-3]) + 4}{time_chek(sheet_per.cell(row=4, column=kol).value)[-3:]}"

                time_end = f"{int(time_chek(sheet_per.cell(row=5, column=kol).value)[:-3]) + 4}{time_chek(sheet_per.cell(row=5, column=kol).value)[-3:]}"

                sheet_list.cell(row=no,
                                column=2).value = f"{time_start}-{time_end}"
            else:
                sheet_list.cell(row=no,
                            column=2).value = f"{time_chek(sheet_per.cell(row=4, column=kol).value)}-{time_chek(sheet_per.cell(row=5, column=kol).value)}"
            per_berin = 10
            per_end = 10
            no_sm = 3
            for st in range(10, sheet_per.max_row + 1):
                if sheet_per.cell(row=st, column=kol).value == 5 and st > per_end:
                    per_berin = st
                    for st_per in range(st, sheet_per.max_row + 1):
                        if sheet_per.cell(row=st_per, column=kol).value != 5:
                            per_end = st_per - 1
                            break
                    # print(per_berin, per_end)
                    for st_beg in range(per_berin, per_berin - 12, -1):
                        if sheet_per.cell(row=st_beg, column=1).value is not None:
                            hour_begin = time_shift(shift(sheet_per.cell(row=st_beg, column=1).value)[0])[0]
                            break

                    min_begin = time_shift(sheet_per.cell(row=per_berin, column=2).value)[0]
                    for st_end in range(per_end, per_end - 12, -1):
                        if sheet_per.cell(row=st_end, column=1).value is not None:
                            hour_end = time_shift(shift(sheet_per.cell(row=st_end, column=1).value)[0])[0]
                            break

                    min_end = time_shift(sheet_per.cell(row=per_end, column=2).value)[1]
                    if min_end == "60":
                        min_end = "00"
                        hour_end = int(hour_end) + 1
                    elif min_end == "0":
                        min_end = "00"
                    elif min_end == "5":
                        min_end = "05"

                    if min_begin == "0":
                        min_begin = "00"
                    elif min_begin == "5":
                        min_begin = "05"
                    elif min_begin == "60":
                        min_begin = "00"
                        hour_begin = int(hour_begin) + 1

                    if CC_name == "НСК":
                        sheet_list.cell(row=no, column=no_sm).value = f"{int(hour_begin) + 4}:{min_begin}-{int(hour_end) + 4}:{min_end}"
                    else:
                        sheet_list.cell(row=no, column=no_sm).value = f"{hour_begin}:{min_begin}-{hour_end}:{min_end}"
                    no_sm += 1
            no += 1
    for i in range(1, sheet_list.max_column + 1):
        for j in range(1, sheet_list.max_row + 1):
            sheet_list.cell(row=j, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    return ()


def get_grafik_oktel(sheet_per, sheet_list):
    # формируем рассылку для CC_name
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    medium = Side(border_style="medium", color="000000")
    no = 2
    for kol in range(4, sheet_per.max_column + 1):
        sheet_list.cell(row=no, column=1).value = sheet_per.cell(row=2, column=kol).value
        CC_name = sheet_per.cell(row=1, column=kol).value
        hour_beg_temp = time_shift(time_chek(sheet_per.cell(row=4, column=kol).value))[0]
        if len(hour_beg_temp) == 1:
            hour_beg_temp = "0" + hour_beg_temp
        #print(hour_beg_temp)
        # hour_beg_temp = sheet_per.cell(row=4, column=kol).value
        # if len(hour_beg_temp) == 4:
        #     hour_beg_temp = "0" + hour_beg_temp
        hour_end_temp = time_shift(time_chek(sheet_per.cell(row=5, column=kol).value))[0]
        if len(hour_end_temp) == 1:
            hour_end_temp = "0" + hour_beg_temp
        # hour_end_temp = sheet_per.cell(row=5, column=kol).value
        # if len(hour_end_temp) == 4:
        #     hour_beg_temp = "0" + hour_beg_temp
        sheet_list.cell(row=no,
                    column=2).value = f"{hour_beg_temp}:" \
                                      f"{time_shift(time_chek(sheet_per.cell(row=4, column=kol).value))[1]}-" \
                                      f"{hour_end_temp}:" \
                                      f"{time_shift(time_chek(sheet_per.cell(row=5, column=kol).value))[1]}"
        #sheet_list.cell(row=no, column=2).value = hour_beg_temp + "-" + hour_end_temp
        #print(sheet_list.cell(row=no, column=2).value)
        per_berin = 10
        per_end = 10
        no_sm = 3
        for st in range(10, sheet_per.max_row + 1):
            if sheet_per.cell(row=st, column=kol).value == 5 and st > per_end:
                per_berin = st
                for st_per in range(st, sheet_per.max_row + 1):
                    if sheet_per.cell(row=st_per, column=kol).value != 5:
                        per_end = st_per - 1
                        break
                # print(per_berin, per_end)
                for st_beg in range(per_berin, per_berin - 12, -1):
                    if sheet_per.cell(row=st_beg, column=1).value is not None:
                        hour_begin = time_shift(shift(sheet_per.cell(row=st_beg, column=1).value)[0])[0]
                        if len(str(hour_begin)) == 1:
                            hour_begin = "0" + str(hour_begin)
                        break

                min_begin = time_shift(sheet_per.cell(row=per_berin, column=2).value)[0]
                for st_end in range(per_end, per_end - 12, -1):
                    if sheet_per.cell(row=st_end, column=1).value is not None:
                        hour_end = time_shift(shift(sheet_per.cell(row=st_end, column=1).value)[0])[0]
                        if len(str(hour_end)) == 1:
                            hour_end = "0" + str(hour_end)
                        break

                min_end = time_shift(sheet_per.cell(row=per_end, column=2).value)[1]
                if min_end == "60":
                    min_end = "00"
                    hour_end = int(hour_end) + 1
                    if len(str(hour_end)) == 1:
                        hour_end = "0" + str(hour_end)
                elif min_end == "0":
                    min_end = "00"
                elif min_end == "5":
                    min_end = "05"

                if min_begin == "0":
                    min_begin = "00"
                elif min_begin == "5":
                    min_begin = "05"
                elif min_begin == "60":
                    min_begin = "00"
                    hour_begin = int(hour_begin) + 1
                    if len(str(hour_begin)) == 1:
                        hour_begin = "0" + str(hour_begin)
                sheet_list.cell(row=no, column=no_sm).value = f"{hour_begin}:{min_begin}-{hour_end}:{min_end}"
                no_sm += 1
        no += 1
    for i in range(1, sheet_list.max_column + 1):
        for j in range(1, sheet_list.max_row + 1):
            sheet_list.cell(row=j, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    return ()