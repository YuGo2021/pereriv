
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side
import pereriv_lib as per


    
wb_grafik1 = openpyxl.load_workbook("Смирновка.xlsx")
sheet_grafik1 = wb_grafik1.active

wb_grafik2 = openpyxl.load_workbook("Высота.xlsx")
sheet_grafik2 = wb_grafik2.active

wb_grafik3 = openpyxl.load_workbook("Киров.xlsx")
sheet_grafik3 = wb_grafik3.active

wb_grafik4 = openpyxl.load_workbook("НСК.xlsx")
sheet_grafik4 = wb_grafik4.active

wb_grafik5 = openpyxl.load_workbook("Ростов.xlsx")
sheet_grafik5 = wb_grafik5.active

wb_grafik6 = openpyxl.load_workbook("Нижний Новгород.xlsx")
sheet_grafik6 = wb_grafik6.active

wb_rez = openpyxl.Workbook()
wb_rez.create_sheet(title = 'перерывы', index = 0)
sheet_rez = wb_rez['перерывы']

redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
greenFill = PatternFill(start_color='1FB714', end_color='1FB714', fill_type='solid')
yellowFill = PatternFill(start_color='FCF305', end_color='FCF305', fill_type='solid')

sheet_rez['A1'] = "Площадка"
sheet_rez['A2'] = "ФИО"
sheet_rez['A3'] = "Смена"
sheet_rez['A4'] = "Начало смены"
sheet_rez['A5'] = "Окончание смены"
sheet_rez['A6'] = "Смена/Доп"
sheet_rez['A7'] = "Рабочие часы из графика"
sheet_rez['A8'] = "Положенный перерыв"
sheet_rez['A9'] = "Сумма перерывов"

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
medium = Side(border_style="medium", color="000000")

i_op = 0
# Собираем Смирновку 
sheet_rez = per.pereriv_CC(sheet_grafik1, i_op, sheet_rez, "Смирновка")[0]
i_op = per.pereriv_CC(sheet_grafik1, i_op, sheet_rez, "Смирновка")[1]

# собираем Высоту
sheet_rez = per.pereriv_CC(sheet_grafik2, i_op, sheet_rez, "Высота")[0]
i_op = per.pereriv_CC(sheet_grafik2, i_op, sheet_rez, "Высота")[1]

# собираем Киров
sheet_rez = per.pereriv_CC(sheet_grafik3, i_op, sheet_rez, "Киров")[0]
i_op = per.pereriv_CC(sheet_grafik3, i_op, sheet_rez, "Киров")[1]

# собираем Новосибирск
sheet_rez = per.pereriv_CC(sheet_grafik4, i_op, sheet_rez, "НСК")[0]
i_op = per.pereriv_CC(sheet_grafik4, i_op, sheet_rez, "НСК")[1]
    
# собираем Ростов
sheet_rez = per.pereriv_CC(sheet_grafik5, i_op, sheet_rez, "Ростов")[0]
i_op = per.pereriv_CC(sheet_grafik5, i_op, sheet_rez, "Ростов")[1]

# собираем Нижний Новгород
sheet_rez = per.pereriv_CC(sheet_grafik6, i_op, sheet_rez, "НиНо")[0]
i_op = per.pereriv_CC(sheet_grafik6, i_op, sheet_rez, "НиНо")[1]
        
for i in range(1, sheet_rez.max_column +1):
    for j in range(1, sheet_rez.max_row + 1):
        sheet_rez.cell(row = j, column = i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        
try:        
    wb_rez.save(f"перерывы_сборка.xlsx")        
    input("Графики собраны. Открываем файл  перерывы_сборка.xlsx и проверяем корректность начала смен и времени перерывов. Нажмите ENTER для продолжения...")
except OSError:
    input("Невозможно сохранить данные. Закройте файл перерывы_сборка.xlsx и запустите программу заново. Нажмите ENTER для продолжения...")