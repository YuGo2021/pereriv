import openpyxl
import pereriv_lib as per
import datetime
import shutil
from openpyxl.styles import Border, Side

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
medium = Side(border_style="medium", color="000000")

wb_grafik_per = openpyxl.load_workbook(f"перерывы_сборка.xlsx")
for s in range(len(wb_grafik_per.sheetnames)):
    if wb_grafik_per.sheetnames[s] == 'перерывы':
        wb_grafik_per.active = s
        break

wb_oper = openpyxl.load_workbook(f"операторы.xlsx")
for s in range(len(wb_oper.sheetnames)):
    if wb_oper.sheetnames[s] == 'операторы':
        wb_oper.active = s
        break


print("Готовим таблицы для рассылки...")
sheet_per = wb_grafik_per.active
sheet_oper = wb_oper.active

sheet_sm = per.new_sheet_oktel(wb_grafik_per, "Лист1")

# формируем рассылку для Октел
per.get_grafik_oktel(sheet_per, sheet_sm)

today = datetime.date.today()
tomorrow = today + datetime.timedelta(days=1)
tomorrow_str = tomorrow.strftime('%Y%m%d')

i_range = []
for i in range(2, sheet_sm.max_row + 1):
    #print(sheet_sm.cell(row=i, column=1).value)
    temp_fio = sheet_sm.cell(row=i, column=1).value
    temp_fio = temp_fio.replace("ё","е")
    temp_fio = temp_fio.lower()
    #print(temp_fio)
    for j in range(2, sheet_oper.max_row + 1):
        #print(per.fio(sheet_oper.cell(row=j, column=1).value))

        temp_fio2 = per.fio(sheet_oper.cell(row=j, column=1).value)
        temp_fio2 = temp_fio2.replace("ё","е")
        temp_fio2 = temp_fio2.lower()

        #print(temp_fio2)
        if temp_fio == temp_fio2 or temp_fio == temp_fio2[:-3] or temp_fio[:-3] == temp_fio2:
            sheet_sm.cell(row=i, column=1).value = sheet_oper.cell(row=j, column=1).value
            #print(per.fio_full(sheet_oper.cell(row=j, column=1).value))
            break
    if temp_fio is None:
        next
    elif temp_fio == sheet_sm.cell(row=i, column=1).value.lower():
        print(f"Ошибка: {temp_fio}")
        i_range.append(i)

for st in reversed(i_range):
    sheet_sm.delete_rows(st)

sh_name = wb_grafik_per.sheetnames
for name in sh_name:
    if name != "Лист1":
        wb_grafik_per.remove(wb_grafik_per[name])

your_answer = "0"
answer = ['y',"Y","N","n"]
while your_answer not in answer:
    your_answer = input(f"Мы делаем перерывы на {tomorrow}? Введите Y/N и нажмите Enter ")

if your_answer == "n" or your_answer == "N":
    while len(answer) != 8:
        answer = input("Введите дату, на которую делаем перерывы в формате ГГГГММДД: ")
    tomorrow_str = answer

try:
    wb_grafik_per.save(f"user{tomorrow_str}.xlsx")
    print(f"Файл перерывов user{tomorrow_str}.xlsx подготовлен.")

except OSError:
    input(f"Невозможно сохранить данные. Закройте файл user{tomorrow_str}.xlsx и "
          "запустите программу заново. Нажмите ENTER для продолжения...")

try:
    wb_next_day = openpyxl.load_workbook(f"\\\\SW-OKTEL-DB-03\\Grafic\\user{tomorrow_str}.xlsx")
    #print(f"Файл ранее был скопирован  в директорию ..\\user{tomorrow_str}.xlsx")
    for s in range(len(wb_next_day.sheetnames)):
        if wb_next_day.sheetnames[s] == 'Лист1':
            wb_next_day.active = s
            break
    sheet_next_day = wb_next_day.active
    for i in range(2, sheet_sm.max_row + 1):
        #print(sheet_sm.cell(row=i, column=1).value)
        if sheet_sm.cell(row=i, column=1).value == None:
            break
        for j in range(2,sheet_next_day.max_row + 1):
            #print(sheet_next_day.cell(row=j, column=1).value)
            if sheet_next_day.cell(row=j, column=1).value == None:
                for z in range(1, 8):
                    sheet_next_day.cell(row=j, column=z).value = sheet_sm.cell(row=i, column=z).value
                break
            elif sheet_sm.cell(row=i, column=1).value == sheet_next_day.cell(row=j, column=1).value:
                for z_col in range(2, 8):
                    sheet_next_day.cell(row=j, column=z_col).value = sheet_sm.cell(row=i, column=z_col).value
                break
            elif j == sheet_next_day.max_row:
                for z in range(1, 8):
                    sheet_next_day.cell(row=j+1, column=z).value = sheet_sm.cell(row=i, column=z).value
                break
            else:
                continue

    for i in range(1, sheet_next_day.max_column + 1):
        for j in range(1, sheet_next_day.max_row + 1):
            sheet_next_day.cell(row=j, column=i).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    try:
        wb_next_day.save(f"\\\\SW-OKTEL-DB-03\\Grafic\\user{tomorrow_str}.xlsx")
        input(f"Файл user{tomorrow_str}.xlsx на сетевом диске \\\\SW-OKTEL-DB-03\Grafic проверен и дополнен вашими данными")
    except OSError:
        input(f"Невозможно сохранить данные. Закройте файл user{tomorrow_str}.xlsx на сетевом диске \\\\SW-OKTEL-DB-03\Grafic и "
              "запустите программу заново. Нажмите ENTER для продолжения...")
    except FileNotFoundError:
        input(f"Невозможно получить доступ к сетевому диску \\\\SW-OKTEL-DB-03\Grafic. Нажмите ENTER для продолжения...")

except OSError:
    try:
        shutil.copyfile(f"user{tomorrow_str}.xlsx", f"\\\\SW-OKTEL-DB-03\\Grafic\\user{tomorrow_str}.xlsx")
        input(f"Файл скопирован  в сетевую директорию \\\\SW-OKTEL-DB-03\\Grafic\\user{tomorrow_str}.xlsx. "
              f"Нажмите ENTER для продолжения...")
    except FileNotFoundError:
        input(
            f"Невозможно получить доступ к сетевому диску \\\\SW-OKTEL-DB-03\Grafic. Нажмите ENTER для продолжения...")


